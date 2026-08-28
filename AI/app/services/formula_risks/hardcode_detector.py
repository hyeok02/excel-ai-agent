from openpyxl.formula.translate import Translator
from openpyxl.worksheet.worksheet import Worksheet

from app.services.formula_risks.finding_factory import build_hardcoded_finding
from app.services.formula_risks.models import FormulaRiskFinding


def detect_hardcoded_values(
    worksheets: list[Worksheet],
) -> list[FormulaRiskFinding]:
    findings: list[FormulaRiskFinding] = []
    for worksheet in worksheets:
        for cell in list(worksheet._cells.values()):
            if not _is_candidate(cell.value, cell.data_type):
                continue
            expected = _expected_formula(worksheet, cell.row, cell.column)
            if expected is None:
                continue
            findings.append(
                build_hardcoded_finding(
                    worksheet.title,
                    cell.coordinate,
                    cell.value,
                    expected,
                )
            )
    return findings


def _is_candidate(value: object, data_type: str) -> bool:
    return data_type != "f" and isinstance(value, (int, float)) and not isinstance(value, bool)


def _expected_formula(
    worksheet: Worksheet,
    row: int,
    column: int,
) -> str | None:
    for before, after in (
        ((row - 1, column), (row + 1, column)),
        ((row, column - 1), (row, column + 1)),
    ):
        if min(*before, *after) < 1:
            continue
        first = worksheet.cell(*before)
        second = worksheet.cell(*after)
        if not (_is_formula(first.value) and _is_formula(second.value)):
            continue
        target = worksheet.cell(row, column).coordinate
        first_expected = _translate(first.value, first.coordinate, target)
        second_expected = _translate(second.value, second.coordinate, target)
        if first_expected and first_expected == second_expected:
            return first_expected
    return None


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _translate(formula: str, origin: str, target: str) -> str | None:
    try:
        return Translator(formula, origin=origin).translate_formula(target)
    except (TypeError, ValueError):
        return None
