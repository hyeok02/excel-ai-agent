from dataclasses import dataclass
import re

from openpyxl.worksheet.worksheet import Worksheet

from app.services.regions.header_features import (
    average_numeric_ratio,
    columns,
    following_rows,
    header_label,
    is_formula,
    is_table_header,
    numeric_ratio,
    style_ratio,
    supported_columns,
)
from app.services.regions.models import RegionBounds
from app.services.regions.utils import is_populated
from app.services.semantic_models import SemanticReason

HEADER_PATTERN = re.compile(
    r"(?:지역|부서|상품|제품|항목|구분|코드|이름|성명|담당|상태|일자|날짜|기간|"
    r"월|분기|연도|년도|금액|매출|비용|수량|합계|비율|증감|실적|계획|"
    r"name|date|amount|total|status|category|code|id)",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class SingleRowHeaderDetection:
    row: int
    confidence: float
    reason: SemanticReason


def detect_single_row_header(
    worksheet: Worksheet,
    bounds: RegionBounds,
) -> SingleRowHeaderDetection | None:
    if bounds.max_row - bounds.min_row < 2 or bounds.max_column == bounds.min_column:
        return None
    candidates = [
        detection
        for row in range(bounds.min_row, min(bounds.max_row - 1, bounds.min_row + 4) + 1)
        if (detection := _score_candidate(worksheet, bounds, row)) is not None
    ]
    if len(candidates) != 1:
        return None
    return candidates[0]


def _score_candidate(
    worksheet: Worksheet,
    bounds: RegionBounds,
    row: int,
) -> SingleRowHeaderDetection | None:
    cells = [worksheet.cell(row=row, column=column) for column in columns(bounds)]
    populated = [cell for cell in cells if is_populated(cell.value)]
    if len(populated) < 2 or any(is_formula(cell.value) for cell in populated):
        return None

    following = following_rows(worksheet, bounds, row)
    if len(following) < 2:
        return None
    width = bounds.max_column - bounds.min_column + 1
    coverage = len(populated) / width
    supported = supported_columns(worksheet, bounds, row, populated)
    labels = [header_label(cell.value) for cell in populated]
    short_ratio = sum(label is not None for label in labels) / len(labels)
    keyword_ratio = sum(bool(label and HEADER_PATTERN.search(label)) for label in labels) / len(labels)
    transition = max(0.0, average_numeric_ratio(following) - numeric_ratio(populated))
    style_contrast = style_ratio(populated) > style_ratio(following[0]) + 0.2
    table_header = is_table_header(worksheet, bounds, row)
    if not (table_header or style_contrast or transition >= 0.3 or keyword_ratio >= 0.3):
        return None

    score = min(
        1.0,
        0.2 * coverage
        + 0.15 * short_ratio
        + 0.2 * supported
        + 0.15
        + 0.15 * min(1.0, transition / 0.6)
        + (0.1 if style_contrast else 0.0)
        + 0.1 * keyword_ratio
        + (0.2 if table_header else 0.0),
    )
    if score < 0.65:
        return None
    evidence = tuple(cell.coordinate for cell in populated[:6])
    return SingleRowHeaderDetection(
        row=row,
        confidence=round(score, 2),
        reason=SemanticReason(
            code="single_row_header",
            message="열 채움률·라벨·서식과 아래 데이터 행의 유형 전환을 종합해 단일 헤더로 판단",
            evidence_cells=evidence,
        ),
    )
