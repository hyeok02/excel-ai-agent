from openpyxl.worksheet.worksheet import Worksheet

from app.services.regions.classification import explicit_role, fallback_role, is_total_row
from app.services.regions.models import RegionBounds
from app.services.regions.utils import row_data_count, row_text_count, sub_bounds
from app.services.semantic_models import SemanticRole


def legacy_table_segments(
    worksheet: Worksheet,
    bounds: RegionBounds,
) -> list[tuple[RegionBounds, SemanticRole]]:
    if bounds.max_row - bounds.min_row < 2 or bounds.max_column == bounds.min_column:
        return []
    data_rows = [
        row
        for row in range(bounds.min_row, bounds.max_row + 1)
        if row_data_count(worksheet, bounds, row) > 0
    ]
    if len(data_rows) < 2:
        return []
    header_end = bounds.min_row - 1
    for row in range(bounds.min_row, min(data_rows) + 1):
        if row_data_count(worksheet, bounds, row) == 0 and row_text_count(
            worksheet, bounds, row
        ):
            header_end = row
        else:
            break
    if header_end < bounds.min_row:
        return []
    total_start = bounds.max_row + 1
    for row in range(bounds.max_row, header_end, -1):
        if is_total_row(worksheet, bounds, row):
            total_start = row
        else:
            break
    segments = [
        (sub_bounds(bounds, bounds.min_row, header_end), SemanticRole.HEADER),
        (sub_bounds(bounds, header_end + 1, total_start - 1), SemanticRole.DATA),
    ]
    if total_start <= bounds.max_row:
        segments.append((sub_bounds(bounds, total_start, bounds.max_row), SemanticRole.TOTAL))
    return segments


def context_row_segments(
    worksheet: Worksheet,
    bounds: RegionBounds,
    sheet_role: str | None,
) -> list[tuple[RegionBounds, SemanticRole]]:
    rows = []
    for row in range(bounds.min_row, bounds.max_row + 1):
        row_bounds = sub_bounds(bounds, row, row)
        role = explicit_role(worksheet, row_bounds, sheet_role) or fallback_role(
            worksheet, row_bounds, sheet_role
        )
        rows.append((row, role))
    context_roles = {
        SemanticRole.TITLE,
        SemanticRole.DESCRIPTION,
        SemanticRole.UNIT,
        SemanticRole.NOTE,
        SemanticRole.INSTRUCTION,
        SemanticRole.WARNING,
        SemanticRole.SOURCE_NOTE,
        SemanticRole.RULE_NOTE,
    }
    if len({role for _, role in rows}) == 1 or any(
        role not in context_roles for _, role in rows
    ):
        return []
    segments = []
    start_row, current_role = rows[0]
    for row, role in rows[1:]:
        if role != current_role:
            segments.append((sub_bounds(bounds, start_row, row - 1), current_role))
            start_row, current_role = row, role
    segments.append((sub_bounds(bounds, start_row, rows[-1][0]), current_role))
    return segments
