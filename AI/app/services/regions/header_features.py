from numbers import Number

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.services.regions.models import RegionBounds
from app.services.regions.utils import is_populated


def following_rows(
    worksheet: Worksheet,
    bounds: RegionBounds,
    row: int,
) -> list[list[object]]:
    results = []
    for row_number in range(row + 1, min(bounds.max_row, row + 3) + 1):
        cells = [worksheet.cell(row=row_number, column=column) for column in columns(bounds)]
        populated = [cell for cell in cells if is_populated(cell.value)]
        if len(populated) >= 2:
            results.append(populated)
    return results


def supported_columns(
    worksheet: Worksheet,
    bounds: RegionBounds,
    row: int,
    populated: list[object],
) -> float:
    supported = sum(
        any(
            is_populated(worksheet.cell(row=next_row, column=cell.column).value)
            for next_row in range(row + 1, min(bounds.max_row, row + 3) + 1)
        )
        for cell in populated
    )
    return supported / len(populated)


def is_table_header(worksheet: Worksheet, bounds: RegionBounds, row: int) -> bool:
    for table in worksheet.tables.values():
        min_column, min_row, max_column, _ = range_boundaries(table.ref)
        if min_row == row and min_column <= bounds.min_column and max_column >= bounds.max_column:
            return True
    return False


def header_label(value: object) -> str | None:
    if value is None or is_formula(value):
        return None
    normalized = str(value).strip()
    return normalized if normalized and len(normalized) <= 60 and "\n" not in normalized else None


def numeric_ratio(cells: list[object]) -> float:
    return sum(is_numeric(getattr(cell, "value", None)) for cell in cells) / len(cells)


def average_numeric_ratio(rows: list[list[object]]) -> float:
    return sum(numeric_ratio(row) for row in rows) / len(rows)


def style_ratio(cells: list[object]) -> float:
    return sum(has_header_style(cell) for cell in cells) / len(cells)


def has_header_style(cell: object) -> bool:
    if isinstance(cell, MergedCell):
        return False
    return bool(
        getattr(getattr(cell, "font", None), "bold", False)
        or getattr(getattr(cell, "fill", None), "fill_type", None)
        or getattr(getattr(cell, "alignment", None), "horizontal", None)
        in {"center", "centerContinuous"}
    )


def is_numeric(value: object) -> bool:
    return (isinstance(value, Number) and not isinstance(value, bool)) or is_formula(value)


def is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def columns(bounds: RegionBounds) -> range:
    return range(bounds.min_column, bounds.max_column + 1)
