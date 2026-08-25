from openpyxl.worksheet.worksheet import Worksheet

from app.services.regions.boundaries import (
    detect_base_bounds,
    merge_documentation_bounds,
)
from app.services.regions.models import CellRegion
from app.services.regions.semantics import segment_and_classify
from app.services.regions.utils import coordinate, populated_coordinates, populated_count


def detect_regions(
    worksheet: Worksheet,
    sheet_role: str | None = None,
) -> list[CellRegion]:
    """Detect logical regions using separators, merged cells and row semantics."""
    populated = populated_coordinates(worksheet)
    if not populated:
        return []

    bounds = detect_base_bounds(worksheet, populated)
    if sheet_role == "documentation":
        bounds = merge_documentation_bounds(worksheet, bounds)

    segmented = [
        region
        for item in bounds
        for region in segment_and_classify(worksheet, item, sheet_role)
    ]
    return [
        CellRegion(
            start_cell=coordinate(item.min_row, item.min_column),
            end_cell=coordinate(item.max_row, item.max_column),
            cell_count=populated_count(worksheet, item),
            semantic=semantic,
        )
        for item, semantic in sorted(
            segmented,
            key=lambda entry: (
                entry[0].min_row,
                entry[0].min_column,
                entry[0].max_row,
                entry[0].max_column,
            ),
        )
    ]
