import re

from openpyxl.worksheet.worksheet import Worksheet

from app.services.regions.models import RegionBounds
from app.services.regions.utils import (
    evidence_cells,
    formula_count,
    intersecting_merged_ranges,
    region_text,
    region_values,
    row_data_count,
    row_text_count,
    style_emphasis_count,
    sub_bounds,
)
from app.services.semantic_models import (
    SemanticClassification,
    SemanticReason,
    SemanticRole,
)


_TITLE_KEYWORDS = ("현황", "보고서", "분석", "검토", "대시보드", "summary", "report")
_UNIT_PATTERN = re.compile(r"(?:단위\s*[:：]|금액\s*단위|달성률\s*단위)", re.IGNORECASE)
_WARNING_PATTERN = re.compile(r"(?:주의|경고|유의|변경\s*시|변경하면|주의사항)", re.IGNORECASE)
_SOURCE_PATTERN = re.compile(r"(?:출처|기준일|작성일|source\s*:)", re.IGNORECASE)
_NOTE_PATTERN = re.compile(r"(?:메모\s*[:：]|비고\s*[:：]|참고\s*[:：]|note\s*:)", re.IGNORECASE)
_RULE_PATTERN = re.compile(r"(?:판단\s*기준|검토\s*기준|적용\s*기준|조건|임계|기준값)", re.IGNORECASE)
_INSTRUCTION_PATTERN = re.compile(r"(?:사용\s*안내|사용법|수정하세요|확인하세요|변경하지\s*마세요)", re.IGNORECASE)
_INPUT_PATTERN = re.compile(r"(?:사용자\s*입력|입력값|가정값|기초\s*데이터)", re.IGNORECASE)
_CALCULATION_PATTERN = re.compile(r"(?:계산\s*결과|계산식|산출|중간\s*계산)", re.IGNORECASE)
_OUTPUT_PATTERN = re.compile(r"(?:최종\s*결과|의사결정\s*요약|판정|결과\s*요약)", re.IGNORECASE)
_TOTAL_PATTERN = re.compile(r"^(?:합계|총계|소계|누계|total)$", re.IGNORECASE)


def segment_and_classify(
    worksheet: Worksheet,
    bounds: RegionBounds,
    sheet_role: str | None,
) -> list[tuple[RegionBounds, SemanticClassification]]:
    explicit = _explicit_role(worksheet, bounds, sheet_role)
    if explicit is not None:
        return [(bounds, _classification(worksheet, bounds, explicit))]

    table_segments = _table_segments(worksheet, bounds)
    if table_segments:
        return [
            (segment, _classification(worksheet, segment, role))
            for segment, role in table_segments
        ]

    row_segments = _context_row_segments(worksheet, bounds, sheet_role)
    if len(row_segments) > 1:
        return [
            (segment, _classification(worksheet, segment, role))
            for segment, role in row_segments
        ]

    role = _fallback_role(worksheet, bounds, sheet_role)
    return [(bounds, _classification(worksheet, bounds, role))]


def _explicit_role(
    worksheet: Worksheet,
    bounds: RegionBounds,
    sheet_role: str | None,
) -> SemanticRole | None:
    text = region_text(worksheet, bounds)
    formulas = formula_count(worksheet, bounds)

    if sheet_role == "system":
        return SemanticRole.SYSTEM_CACHE
    if sheet_role == "documentation":
        return SemanticRole.INSTRUCTION
    if _WARNING_PATTERN.search(text):
        return SemanticRole.WARNING
    if _SOURCE_PATTERN.search(text):
        return SemanticRole.SOURCE_NOTE
    if _NOTE_PATTERN.search(text):
        return SemanticRole.NOTE
    if _INPUT_PATTERN.search(text):
        return SemanticRole.INPUT
    if _CALCULATION_PATTERN.search(text):
        return SemanticRole.CALCULATION
    if _OUTPUT_PATTERN.search(text) and formulas:
        return SemanticRole.OUTPUT
    if _RULE_PATTERN.search(text) and formulas == 0:
        return SemanticRole.RULE_NOTE
    return None


def _table_segments(
    worksheet: Worksheet,
    bounds: RegionBounds,
) -> list[tuple[RegionBounds, SemanticRole]]:
    if bounds.max_row - bounds.min_row < 2 or bounds.max_column == bounds.min_column:
        return []

    data_rows = [
        row
        for row in range(bounds.min_row, bounds.max_row + 1)
        if row_data_count(worksheet, bounds, row) > 0
    ]
    if len(data_rows) < 2:
        return []

    header_end = bounds.min_row - 1
    for row in range(bounds.min_row, min(data_rows) + 1):
        if row_data_count(worksheet, bounds, row) == 0 and row_text_count(
            worksheet, bounds, row
        ):
            header_end = row
        else:
            break
    if header_end < bounds.min_row:
        return []

    total_start = bounds.max_row + 1
    for row in range(bounds.max_row, header_end, -1):
        if _is_total_row(worksheet, bounds, row):
            total_start = row
        else:
            break

    data_start = header_end + 1
    data_end = total_start - 1
    if data_start > data_end:
        return []

    segments = [
        (sub_bounds(bounds, bounds.min_row, header_end), SemanticRole.HEADER),
        (sub_bounds(bounds, data_start, data_end), SemanticRole.DATA),
    ]
    if total_start <= bounds.max_row:
        segments.append(
            (sub_bounds(bounds, total_start, bounds.max_row), SemanticRole.TOTAL)
        )
    return segments


def _context_row_segments(
    worksheet: Worksheet,
    bounds: RegionBounds,
    sheet_role: str | None,
) -> list[tuple[RegionBounds, SemanticRole]]:
    rows: list[tuple[int, SemanticRole]] = []
    for row in range(bounds.min_row, bounds.max_row + 1):
        row_bounds = sub_bounds(bounds, row, row)
        role = _explicit_role(worksheet, row_bounds, sheet_role)
        if role is None:
            role = _fallback_role(worksheet, row_bounds, sheet_role)
        rows.append((row, role))

    context_roles = {
        SemanticRole.TITLE,
        SemanticRole.DESCRIPTION,
        SemanticRole.UNIT,
        SemanticRole.NOTE,
        SemanticRole.INSTRUCTION,
        SemanticRole.WARNING,
        SemanticRole.SOURCE_NOTE,
        SemanticRole.RULE_NOTE,
    }
    if len({role for _, role in rows}) == 1 or any(
        role not in context_roles for _, role in rows
    ):
        return []

    segments: list[tuple[RegionBounds, SemanticRole]] = []
    start_row, current_role = rows[0]
    for row, role in rows[1:]:
        if role != current_role:
            segments.append((sub_bounds(bounds, start_row, row - 1), current_role))
            start_row, current_role = row, role
    segments.append((sub_bounds(bounds, start_row, rows[-1][0]), current_role))
    return segments


def _fallback_role(
    worksheet: Worksheet,
    bounds: RegionBounds,
    sheet_role: str | None,
) -> SemanticRole:
    text = region_text(worksheet, bounds)
    values = region_values(worksheet, bounds)
    formulas = formula_count(worksheet, bounds)
    merged = intersecting_merged_ranges(worksheet, bounds)
    emphasized = style_emphasis_count(worksheet, bounds)

    if _UNIT_PATTERN.search(text):
        return SemanticRole.UNIT
    if _INSTRUCTION_PATTERN.search(text):
        return SemanticRole.INSTRUCTION
    if (
        bounds.min_row <= 2
        and len(values) <= 2
        and merged
        and emphasized > 0
        and (
            bounds.min_row == 1
            or any(keyword in text.lower() for keyword in _TITLE_KEYWORDS)
        )
    ):
        return SemanticRole.TITLE
    if formulas:
        return SemanticRole.CALCULATION
    if values and all(isinstance(value, str) for value in values):
        return SemanticRole.DESCRIPTION
    if sheet_role == "input":
        return SemanticRole.INPUT
    if sheet_role == "output":
        return SemanticRole.OUTPUT
    return SemanticRole.DATA


def _classification(
    worksheet: Worksheet,
    bounds: RegionBounds,
    role: SemanticRole,
) -> SemanticClassification:
    reason_code, message = _ROLE_REASONS.get(
        role,
        ("semantic_fallback", "셀 값과 서식 분포를 기준으로 역할을 판단"),
    )
    reasons = (
        SemanticReason(
            code=reason_code,
            message=message,
            evidence_cells=evidence_cells(worksheet, bounds),
        ),
        *bounds.boundary_reasons,
    )
    return SemanticClassification(
        role=role,
        confidence=_ROLE_CONFIDENCE.get(role, 0.72),
        reasons=reasons,
    )


def _is_total_row(worksheet: Worksheet, bounds: RegionBounds, row: int) -> bool:
    values = region_values(worksheet, sub_bounds(bounds, row, row))
    has_total_label = any(
        isinstance(value, str) and _TOTAL_PATTERN.fullmatch(value.strip())
        for value in values
    )
    return has_total_label and row_data_count(worksheet, bounds, row) > 0


_ROLE_REASONS = {
    SemanticRole.TITLE: ("title_style", "상단 병합·강조 서식과 제목 문구를 탐지"),
    SemanticRole.DESCRIPTION: ("narrative_text", "계산값이 없는 문장형 설명 영역을 탐지"),
    SemanticRole.UNIT: ("unit_label", "단위 표기 문구를 탐지"),
    SemanticRole.HEADER: ("header_style_transition", "데이터 앞의 문자·강조 행을 헤더로 분리"),
    SemanticRole.DATA: ("tabular_data", "헤더 아래 반복되는 데이터 행을 탐지"),
    SemanticRole.TOTAL: ("total_formula_pattern", "합계 문구와 집계 수식 행을 탐지"),
    SemanticRole.INPUT: ("input_heading", "입력·가정값 문구와 상수 값 영역을 탐지"),
    SemanticRole.CALCULATION: ("formula_distribution", "계산 문구와 수식 분포를 탐지"),
    SemanticRole.OUTPUT: ("output_heading", "판정·결과 문구와 결과 수식을 탐지"),
    SemanticRole.INSTRUCTION: ("instruction_text", "사용 방법과 작업 순서를 설명하는 문장을 탐지"),
    SemanticRole.WARNING: ("warning_text", "주의·경고 표현이 포함된 문장을 탐지"),
    SemanticRole.SOURCE_NOTE: ("source_note_text", "출처·기준일 문구를 탐지"),
    SemanticRole.NOTE: ("note_text", "메모·비고·참고 문구를 탐지"),
    SemanticRole.RULE_NOTE: ("rule_note_text", "판단·검토 기준을 설명하는 문장을 탐지"),
    SemanticRole.SYSTEM_CACHE: ("system_policy", "시스템 캐시 시트 정책을 적용"),
}

_ROLE_CONFIDENCE = {
    SemanticRole.TITLE: 0.94,
    SemanticRole.UNIT: 0.97,
    SemanticRole.WARNING: 0.96,
    SemanticRole.SOURCE_NOTE: 0.96,
    SemanticRole.NOTE: 0.92,
    SemanticRole.RULE_NOTE: 0.9,
    SemanticRole.INSTRUCTION: 0.9,
    SemanticRole.HEADER: 0.88,
    SemanticRole.TOTAL: 0.94,
    SemanticRole.INPUT: 0.9,
    SemanticRole.CALCULATION: 0.92,
    SemanticRole.OUTPUT: 0.92,
    SemanticRole.DATA: 0.84,
    SemanticRole.SYSTEM_CACHE: 1.0,
}
