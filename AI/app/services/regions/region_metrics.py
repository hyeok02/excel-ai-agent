from numbers import Number

from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from app.services.regions.models import RegionBounds
from app.services.regions.region_cells import region_values, sub_bounds


def row_text_count(worksheet: Worksheet, bounds: RegionBounds, row: int) -> int:
    return sum(
        isinstance(value, str) and not value.startswith("=")
        for value in region_values(worksheet, sub_bounds(bounds, row, row))
    )


def row_data_count(worksheet: Worksheet, bounds: RegionBounds, row: int) -> int:
    return sum(
        isinstance(value, Number)
        or (isinstance(value, str) and value.startswith("="))
        for value in region_values(worksheet, sub_bounds(bounds, row, row))
    )


def formula_count(worksheet: Worksheet, bounds: RegionBounds) -> int:
    return sum(
        1
        for row in range(bounds.min_row, bounds.max_row + 1)
        for column in range(bounds.min_column, bounds.max_column + 1)
        if getattr(worksheet.cell(row=row, column=column), "data_type", None) == "f"
    )


def style_emphasis_count(worksheet: Worksheet, bounds: RegionBounds) -> int:
    return sum(
        _has_structural_style(worksheet.cell(row=row, column=column))
        for row in range(bounds.min_row, bounds.max_row + 1)
        for column in range(bounds.min_column, bounds.max_column + 1)
        if not isinstance(worksheet.cell(row=row, column=column), MergedCell)
    )


def _has_structural_style(cell: object) -> bool:
    font = getattr(cell, "font", None)
    fill = getattr(cell, "fill", None)
    border = getattr(cell, "border", None)
    alignment = getattr(cell, "alignment", None)
    border_sides = (
        getattr(border, side, None) for side in ("left", "right", "top", "bottom")
    )
    return bool(
        getattr(font, "bold", False)
        or getattr(fill, "fill_type", None)
        or any(getattr(side, "style", None) for side in border_sides)
        or getattr(alignment, "horizontal", None) in {"center", "centerContinuous"}
    )
