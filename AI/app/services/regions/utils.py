from numbers import Number

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.regions.models import RegionBounds


def sub_bounds(
    bounds: RegionBounds,
    min_row: int,
    max_row: int,
) -> RegionBounds:
    return RegionBounds(
        min_row=min_row,
        max_row=max_row,
        min_column=bounds.min_column,
        max_column=bounds.max_column,
        boundary_reasons=bounds.boundary_reasons,
    )


def region_values(worksheet: Worksheet, bounds: RegionBounds) -> list[object]:
    values = []
    seen_merged_anchors: set[tuple[int, int]] = set()
    for row in range(bounds.min_row, bounds.max_row + 1):
        for column in range(bounds.min_column, bounds.max_column + 1):
            cell = worksheet.cell(row=row, column=column)
            value = cell.value
            if isinstance(cell, MergedCell):
                anchor = merged_anchor(worksheet, row, column)
                if anchor is None or anchor in seen_merged_anchors:
                    continue
                seen_merged_anchors.add(anchor)
                value = worksheet.cell(row=anchor[0], column=anchor[1]).value
            if is_populated(value):
                values.append(value)
    return values


def region_text(worksheet: Worksheet, bounds: RegionBounds) -> str:
    return " ".join(
        str(value).strip()
        for value in region_values(worksheet, bounds)
        if isinstance(value, str) and not value.startswith("=")
    )


def row_text_count(worksheet: Worksheet, bounds: RegionBounds, row: int) -> int:
    return sum(
        isinstance(value, str) and not value.startswith("=")
        for value in region_values(worksheet, sub_bounds(bounds, row, row))
    )


def row_data_count(worksheet: Worksheet, bounds: RegionBounds, row: int) -> int:
    return sum(
        isinstance(value, Number)
        or (isinstance(value, str) and value.startswith("="))
        for value in region_values(worksheet, sub_bounds(bounds, row, row))
    )


def formula_count(worksheet: Worksheet, bounds: RegionBounds) -> int:
    return sum(
        1
        for row in range(bounds.min_row, bounds.max_row + 1)
        for column in range(bounds.min_column, bounds.max_column + 1)
        if getattr(worksheet.cell(row=row, column=column), "data_type", None) == "f"
    )


def style_emphasis_count(worksheet: Worksheet, bounds: RegionBounds) -> int:
    return sum(
        _has_structural_style(worksheet.cell(row=row, column=column))
        for row in range(bounds.min_row, bounds.max_row + 1)
        for column in range(bounds.min_column, bounds.max_column + 1)
        if not isinstance(worksheet.cell(row=row, column=column), MergedCell)
    )


def _has_structural_style(cell: object) -> bool:
    font = getattr(cell, "font", None)
    fill = getattr(cell, "fill", None)
    border = getattr(cell, "border", None)
    alignment = getattr(cell, "alignment", None)
    border_sides = (
        getattr(border, side, None)
        for side in ("left", "right", "top", "bottom")
    )
    return bool(
        getattr(font, "bold", False)
        or getattr(fill, "fill_type", None)
        or any(getattr(side, "style", None) for side in border_sides)
        or getattr(alignment, "horizontal", None) in {"center", "centerContinuous"}
    )


def intersecting_merged_ranges(
    worksheet: Worksheet,
    bounds: RegionBounds,
) -> list[str]:
    return [
        str(merged_range)
        for merged_range in worksheet.merged_cells.ranges
        if merged_range.max_row >= bounds.min_row
        and merged_range.min_row <= bounds.max_row
        and merged_range.max_col >= bounds.min_column
        and merged_range.min_col <= bounds.max_column
    ]


def merged_anchor(
    worksheet: Worksheet,
    row: int,
    column: int,
) -> tuple[int, int] | None:
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= column <= merged_range.max_col
        ):
            return merged_range.min_row, merged_range.min_col
    return None


def populated_coordinates(worksheet: Worksheet) -> set[tuple[int, int]]:
    return {
        (cell.row, cell.column)
        for cell in worksheet._cells.values()
        if is_populated(cell.value)
    }


def populated_count(worksheet: Worksheet, bounds: RegionBounds) -> int:
    return sum(
        1
        for row, column in populated_coordinates(worksheet)
        if bounds.min_row <= row <= bounds.max_row
        and bounds.min_column <= column <= bounds.max_column
    )


def evidence_cells(
    worksheet: Worksheet,
    bounds: RegionBounds,
) -> tuple[str, ...]:
    cells = [
        coordinate(row, column)
        for row, column in sorted(populated_coordinates(worksheet))
        if bounds.min_row <= row <= bounds.max_row
        and bounds.min_column <= column <= bounds.max_column
    ]
    if not cells:
        cells = [
            coordinate(bounds.min_row, bounds.min_column),
            coordinate(bounds.max_row, bounds.max_column),
        ]
    return tuple(dict.fromkeys(cells[:6]))


def is_populated(value: object) -> bool:
    return value is not None and not (isinstance(value, str) and not value.strip())


def coordinate(row: int, column: int) -> str:
    return f"{get_column_letter(column)}{row}"
