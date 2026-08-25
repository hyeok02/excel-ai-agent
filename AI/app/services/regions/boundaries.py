from openpyxl.worksheet.worksheet import Worksheet

from app.services.regions.models import RegionBounds
from app.services.regions.utils import (
    coordinate,
    formula_count,
    intersecting_merged_ranges,
    is_populated,
    populated_coordinates,
)
from app.services.semantic_models import SemanticReason


def detect_base_bounds(
    worksheet: Worksheet,
    populated: set[tuple[int, int]],
) -> list[RegionBounds]:
    occupied = set(populated)
    for merged_range in worksheet.merged_cells.ranges:
        if (merged_range.min_row, merged_range.min_col) not in populated:
            continue
        occupied.update(
            (row, column)
            for row in range(merged_range.min_row, merged_range.max_row + 1)
            for column in range(merged_range.min_col, merged_range.max_col + 1)
        )

    row_groups = _contiguous_groups({row for row, _ in occupied})
    results: list[RegionBounds] = []
    for min_row, max_row in row_groups:
        column_groups = _column_groups(occupied, min_row, max_row)
        shared_heading = _shared_merged_heading(
            worksheet,
            min_row,
            max_row,
            column_groups,
        )
        if shared_heading is not None:
            column_groups = _column_groups(occupied, min_row + 1, max_row)

        for min_column, max_column in column_groups:
            group_rows = [
                row
                for row, column in occupied
                if min_column <= column <= max_column
                and min_row <= row <= max_row
            ]
            group_min_row = min(group_rows)
            group_max_row = max(group_rows)
            reasons = _boundary_reasons(
                group_min_row,
                group_max_row,
                min_column,
                max_column,
                len(column_groups),
                shared_heading,
            )
            results.append(
                RegionBounds(
                    min_row=group_min_row,
                    max_row=group_max_row,
                    min_column=min_column,
                    max_column=max_column,
                    boundary_reasons=reasons,
                )
            )
    return _split_isolated_bounds(worksheet, results)


def merge_documentation_bounds(
    worksheet: Worksheet,
    bounds: list[RegionBounds],
) -> list[RegionBounds]:
    if len(bounds) < 2:
        return bounds
    merged: list[RegionBounds] = []
    current = bounds[0]
    for candidate in bounds[1:]:
        gap = candidate.min_row - current.max_row - 1
        overlaps = not (
            candidate.max_column < current.min_column
            or candidate.min_column > current.max_column
        )
        combined = RegionBounds(
            min_row=current.min_row,
            max_row=candidate.max_row,
            min_column=min(current.min_column, candidate.min_column),
            max_column=max(current.max_column, candidate.max_column),
            boundary_reasons=(
                *current.boundary_reasons,
                SemanticReason(
                    code="documentation_flow",
                    message="한 줄 간격의 제목과 안내 본문을 하나의 설명 흐름으로 결합",
                    evidence_cells=(
                        coordinate(current.min_row, current.min_column),
                        coordinate(candidate.max_row, candidate.max_column),
                    ),
                ),
            ),
        )
        if gap <= 1 and overlaps and formula_count(worksheet, combined) == 0:
            current = combined
        else:
            merged.append(current)
            current = candidate
    merged.append(current)
    return merged


def _boundary_reasons(
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
    group_count: int,
    shared_heading: str | None,
) -> tuple[SemanticReason, ...]:
    reasons = []
    evidence = tuple(
        dict.fromkeys(
            (
                coordinate(min_row, min_column),
                coordinate(max_row, max_column),
            )
        )
    )
    if min_row > 1:
        reasons.append(
            SemanticReason(
                code="blank_row_boundary",
                message="빈 행을 기준으로 위아래 영역을 분리",
                evidence_cells=evidence,
            )
        )
    if group_count > 1:
        reasons.append(
            SemanticReason(
                code="blank_column_boundary",
                message="빈 열을 기준으로 좌우 영역을 분리",
                evidence_cells=evidence,
            )
        )
    if shared_heading is not None:
        reasons.append(
            SemanticReason(
                code="shared_merged_heading",
                message="공통 병합 제목 아래의 열 묶음을 독립 영역으로 분리",
                evidence_cells=(shared_heading,),
            )
        )
    return tuple(reasons)


def _shared_merged_heading(
    worksheet: Worksheet,
    min_row: int,
    max_row: int,
    column_groups: list[tuple[int, int]],
) -> str | None:
    if min_row == max_row or len(column_groups) != 1:
        return None
    lower_occupied = {
        (cell.row, cell.column)
        for cell in worksheet._cells.values()
        if min_row < cell.row <= max_row and is_populated(cell.value)
    }
    lower_groups = _column_groups(lower_occupied, min_row + 1, max_row)
    if len(lower_groups) <= 1:
        return None
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_row == min_row and merged_range.max_row == min_row:
            return str(merged_range)
    return None


def _split_isolated_bounds(
    worksheet: Worksheet,
    bounds: list[RegionBounds],
) -> list[RegionBounds]:
    results: list[RegionBounds] = []
    populated = populated_coordinates(worksheet)
    for item in bounds:
        coordinates = {
            cell
            for cell in populated
            if item.min_row <= cell[0] <= item.max_row
            and item.min_column <= cell[1] <= item.max_column
        }
        has_adjacent_cells = any(
            (row + row_offset, column + column_offset) in coordinates
            for row, column in coordinates
            for row_offset, column_offset in ((1, 0), (0, 1))
        )
        if (
            len(coordinates) > 1
            and not has_adjacent_cells
            and not intersecting_merged_ranges(worksheet, item)
        ):
            results.extend(
                RegionBounds(
                    min_row=row,
                    max_row=row,
                    min_column=column,
                    max_column=column,
                    boundary_reasons=item.boundary_reasons,
                )
                for row, column in sorted(coordinates)
            )
        else:
            results.append(item)
    return results


def _column_groups(
    occupied: set[tuple[int, int]],
    min_row: int,
    max_row: int,
) -> list[tuple[int, int]]:
    columns = {column for row, column in occupied if min_row <= row <= max_row}
    return _contiguous_groups(columns)


def _contiguous_groups(values: set[int]) -> list[tuple[int, int]]:
    if not values:
        return []
    sorted_values = sorted(values)
    groups: list[tuple[int, int]] = []
    start = previous = sorted_values[0]
    for value in sorted_values[1:]:
        if value > previous + 1:
            groups.append((start, previous))
            start = value
        previous = value
    groups.append((start, previous))
    return groups
