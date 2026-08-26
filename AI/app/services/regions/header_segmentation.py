from openpyxl.worksheet.worksheet import Worksheet

from app.services.regions.classification import (
    classify,
    explicit_role,
    fallback_role,
    is_total_row,
)
from app.services.regions.models import RegionBounds
from app.services.regions.table_segmentation import context_row_segments
from app.services.regions.utils import sub_bounds
from app.services.semantic_models import SemanticClassification, SemanticReason, SemanticRole


def header_segments(
    worksheet: Worksheet,
    bounds: RegionBounds,
    sheet_role: str | None,
    start_row: int,
    end_row: int,
    confidence: float,
    reason: SemanticReason,
) -> list[tuple[RegionBounds, SemanticClassification]]:
    results = _leading_segments(worksheet, bounds, sheet_role, start_row)
    header_bounds = sub_bounds(bounds, start_row, end_row)
    results.append(
        (
            header_bounds,
            classify(
                worksheet,
                header_bounds,
                SemanticRole.HEADER,
                confidence=confidence,
                extra_reasons=(reason,),
            ),
        )
    )
    total_start = _total_start(worksheet, bounds, end_row)
    if end_row + 1 < total_start:
        data_bounds = sub_bounds(bounds, end_row + 1, total_start - 1)
        results.append((data_bounds, classify(worksheet, data_bounds, SemanticRole.DATA)))
    if total_start <= bounds.max_row:
        total_bounds = sub_bounds(bounds, total_start, bounds.max_row)
        results.append((total_bounds, classify(worksheet, total_bounds, SemanticRole.TOTAL)))
    return results


def _total_start(worksheet: Worksheet, bounds: RegionBounds, header_end: int) -> int:
    total_start = bounds.max_row + 1
    for row in range(bounds.max_row, header_end, -1):
        if not is_total_row(worksheet, bounds, row):
            break
        total_start = row
    return total_start


def _leading_segments(
    worksheet: Worksheet,
    bounds: RegionBounds,
    sheet_role: str | None,
    header_row: int,
) -> list[tuple[RegionBounds, SemanticClassification]]:
    if header_row == bounds.min_row:
        return []
    leading = sub_bounds(bounds, bounds.min_row, header_row - 1)
    segments = context_row_segments(worksheet, leading, sheet_role)
    if not segments:
        role = explicit_role(worksheet, leading, sheet_role) or fallback_role(
            worksheet, leading, sheet_role
        )
        segments = [(leading, role)]
    return [(segment, classify(worksheet, segment, role)) for segment, role in segments]
