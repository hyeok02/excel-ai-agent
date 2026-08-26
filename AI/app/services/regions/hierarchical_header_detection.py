from dataclasses import dataclass

from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from app.services.regions.models import RegionBounds
from app.services.regions.utils import (
    evidence_cells,
    row_data_count,
    row_text_count,
    sub_bounds,
)
from app.services.semantic_models import SemanticReason

MAX_HEADER_ROWS = 4


@dataclass(frozen=True)
class HierarchicalHeaderDetection:
    start_row: int
    end_row: int
    confidence: float
    reason: SemanticReason


def detect_merged_hierarchical_header(
    worksheet: Worksheet,
    bounds: RegionBounds,
) -> HierarchicalHeaderDetection | None:
    if bounds.max_row - bounds.min_row < 3 or bounds.max_column == bounds.min_column:
        return None
    header_end = _header_end(worksheet, bounds)
    if header_end is None or header_end == bounds.min_row:
        return None
    header_bounds = sub_bounds(bounds, bounds.min_row, header_end)
    if not _has_hierarchical_merge(worksheet, header_bounds):
        return None
    if not _has_tabular_rows_below(worksheet, bounds, header_end):
        return None
    confidence = min(0.96, 0.78 + 0.04 * (header_end - bounds.min_row))
    return HierarchicalHeaderDetection(
        start_row=bounds.min_row,
        end_row=header_end,
        confidence=round(confidence, 2),
        reason=SemanticReason(
            code="merged_hierarchical_header",
            message="병합된 상위 항목과 하위 열 이름의 배치를 계층형 헤더로 판단",
            evidence_cells=evidence_cells(worksheet, header_bounds),
        ),
    )


def _header_end(worksheet: Worksheet, bounds: RegionBounds) -> int | None:
    scan_end = min(bounds.max_row - 2, bounds.min_row + MAX_HEADER_ROWS - 1)
    header_end = None
    for row in range(bounds.min_row, scan_end + 1):
        if row_data_count(worksheet, bounds, row) > 0:
            break
        if row_text_count(worksheet, bounds, row) == 0:
            break
        header_end = row
    return header_end


def _has_hierarchical_merge(
    worksheet: Worksheet,
    bounds: RegionBounds,
) -> bool:
    width = bounds.max_column - bounds.min_column + 1
    for merged in worksheet.merged_cells.ranges:
        if not _intersects(merged, bounds):
            continue
        merged_width = merged.max_col - merged.min_col + 1
        merged_height = merged.max_row - merged.min_row + 1
        if merged_height > 1 or 1 < merged_width < width:
            return True
    return False


def _intersects(merged: CellRange, bounds: RegionBounds) -> bool:
    return bool(
        merged.max_row >= bounds.min_row
        and merged.min_row <= bounds.max_row
        and merged.max_col >= bounds.min_column
        and merged.min_col <= bounds.max_column
    )


def _has_tabular_rows_below(
    worksheet: Worksheet,
    bounds: RegionBounds,
    header_end: int,
) -> bool:
    following = range(header_end + 1, min(bounds.max_row, header_end + 3) + 1)
    return sum(row_data_count(worksheet, bounds, row) > 0 for row in following) >= 2
