from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.services.dependency_analyzer import analyze_dependencies
from app.services.formula_analyzer import analyze_formulas
from app.services.formula_risks import detect_formula_risks
from app.services.sheet_classifier import classify_sheets
from app.services.workbook_parsing.models import (
    ExcludedSheetSummary,
    SheetSummary,
    WorkbookSummary,
)
from app.services.workbook_parsing.sheets import build_sheet_summaries
from app.services.worksheet_filter import evaluate_worksheet_inclusion

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}


class InvalidWorkbookError(ValueError):
    """Raised when an uploaded file cannot be parsed as a supported workbook."""


def parse_workbook(filename: str, content: bytes) -> WorkbookSummary:
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise InvalidWorkbookError(".xlsx 또는 .xlsm 파일만 업로드할 수 있습니다.")
    workbook, value_workbook = _load_workbooks(content, extension == ".xlsm")
    try:
        total_sheet_count = len(workbook.sheetnames)
        inclusions = {
            sheet.title: evaluate_worksheet_inclusion(sheet)
            for sheet in workbook.worksheets
        }
        formulas = {
            sheet.title: analyze_formulas(sheet, value_workbook[sheet.title])
            for sheet in workbook.worksheets
        }
        classifications = classify_sheets(workbook, formulas, inclusions)
        sheets, excluded = build_sheet_summaries(
            workbook,
            value_workbook,
            formulas,
            inclusions,
            classifications,
        )
        formula_risks = detect_formula_risks(
            workbook.sheetnames,
            [(sheet.name, sheet.formulas) for sheet in sheets],
        )
    finally:
        workbook.close()
        value_workbook.close()
    dependencies = analyze_dependencies(
        [(sheet.name, sheet.formulas) for sheet in sheets]
    )
    return WorkbookSummary(
        filename=filename,
        sheet_count=len(sheets),
        sheets=sheets,
        total_sheet_count=total_sheet_count,
        excluded_sheet_count=len(excluded),
        excluded_sheets=excluded,
        dependency_summary=dependencies,
        formula_risk_summary=formula_risks,
    )


def _load_workbooks(content: bytes, keep_vba: bool) -> tuple[object, object]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, keep_vba=keep_vba)
        values = load_workbook(BytesIO(content), data_only=True, keep_vba=keep_vba)
        return workbook, values
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exception:
        raise InvalidWorkbookError("올바른 Excel 파일이 아닙니다.") from exception


__all__ = [
    "ExcludedSheetSummary",
    "InvalidWorkbookError",
    "SheetSummary",
    "WorkbookSummary",
    "parse_workbook",
]
