"""Read Excel compatibility styles without ever changing the source package.

Some exporters wrap style slots in Markup Compatibility AlternateContent.
openpyxl skips those slots instead of selecting their standard Fallback, which
shifts style indices. Only the temporary reading view resolves these wrappers;
writeback must continue to patch and verify the original bytes.
"""

from copy import copy
from io import BytesIO
from xml.etree.ElementTree import ParseError, tostring
from zipfile import BadZipFile, ZipFile

from defusedxml.ElementTree import fromstring
from defusedxml.common import DefusedXmlException
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
STYLES_PATH = "xl/styles.xml"
STYLE_ERROR = (
    "Excel 서식 정보를 읽을 수 없습니다. "
    "파일을 Excel에서 다시 저장한 뒤 업로드해 주세요."
)
COMPATIBILITY_ERROR = (
    "Excel 호환 서식을 읽을 수 없습니다. "
    "파일을 Excel에서 다시 저장한 뒤 업로드해 주세요."
)
STYLE_SLOTS = {
    "cellXfs": "xf", "cellStyleXfs": "xf", "fonts": "font",
    "fills": "fill", "borders": "border", "numFmts": "numFmt",
    "cellStyles": "cellStyle", "dxfs": "dxf",
}


class InvalidWorkbookError(ValueError):
    """A user-readable error for an unsupported or unreadable workbook."""


def prepare_workbook_for_reading(content: bytes) -> bytes:
    """Return the same bytes for normal files, a private reading view otherwise."""
    try:
        with ZipFile(BytesIO(content)) as source:
            if STYLES_PATH not in source.namelist():
                return content
            styles = source.read(STYLES_PATH)
            if not any(
                "AlternateContent".encode(encoding) in styles
                for encoding in ("utf-8", "utf-16-le", "utf-16-be")
            ):
                return content
            root = fromstring(styles, forbid_dtd=True)
            if not _resolve_style_fallbacks(root):
                return content
            output = BytesIO()
            with ZipFile(output, "w") as target:
                target.comment = source.comment
                for info in source.infolist():
                    data = (
                        tostring(root, encoding="utf-8", xml_declaration=True)
                        if info.filename == STYLES_PATH else source.read(info.filename)
                    )
                    # ZipFile.writestr mutates ZipInfo offsets; don't mutate source.
                    target.writestr(copy(info), data)
            return output.getvalue()
    except InvalidWorkbookError:
        raise
    except (ParseError, DefusedXmlException) as exception:
        raise InvalidWorkbookError(COMPATIBILITY_ERROR) from exception
    except (BadZipFile, OSError, KeyError, ValueError, RuntimeError) as exception:
        raise InvalidWorkbookError("올바른 Excel 파일이 아닙니다.") from exception


def _resolve_style_fallbacks(root) -> bool:
    changed = False
    pending = [root]
    while pending:
        parent = pending.pop()
        for index, child in enumerate(list(parent)):
            if child.tag != f"{{{MC_NS}}}AlternateContent":
                continue
            slot = next(
                (value for key, value in STYLE_SLOTS.items()
                 if parent.tag == f"{{{MAIN_NS}}}{key}"),
                None,
            )
            fallbacks = child.findall(f"{{{MC_NS}}}Fallback")
            if (
                slot is None or len(fallbacks) != 1 or len(fallbacks[0]) != 1
                or fallbacks[0][0].tag != f"{{{MAIN_NS}}}{slot}"
            ):
                # Never drop an unresolved slot or shift the remaining style IDs.
                raise InvalidWorkbookError(COMPATIBILITY_ERROR)
            parent.remove(child)
            parent.insert(index, fallbacks[0][0])
            changed = True
        pending.extend(parent)
    return changed


def _load_prepared(content: bytes, *, data_only: bool, keep_vba: bool, read_only: bool):
    try:
        return load_workbook(
            BytesIO(content), data_only=data_only, keep_vba=keep_vba,
            read_only=read_only,
        )
    except (IndexError, TypeError, AttributeError) as exception:
        raise InvalidWorkbookError(STYLE_ERROR) from exception
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError,
            SyntaxError, DefusedXmlException) as exception:
        raise InvalidWorkbookError("올바른 Excel 파일이 아닙니다.") from exception


def load_workbook_for_reading(
    content: bytes, *, data_only: bool = False, keep_vba: bool = False,
    read_only: bool = False,
):
    return _load_prepared(
        prepare_workbook_for_reading(content), data_only=data_only,
        keep_vba=keep_vba, read_only=read_only,
    )


def load_workbook_pair(
    content: bytes, *, keep_vba: bool = False, read_only: bool = False,
) -> tuple[object, object]:
    prepared = prepare_workbook_for_reading(content)
    formulas = _load_prepared(
        prepared, data_only=False, keep_vba=keep_vba, read_only=read_only,
    )
    try:
        values = _load_prepared(
            prepared, data_only=True, keep_vba=keep_vba, read_only=read_only,
        )
    except Exception:
        close_workbook(formulas)
        raise
    return formulas, values


def close_workbook(workbook) -> None:
    """Close the retained VBA archive as well as read-only worksheet handles."""
    try:
        workbook.close()
    finally:
        archive = getattr(workbook, "vba_archive", None)
        if archive is not None:
            archive.close()
