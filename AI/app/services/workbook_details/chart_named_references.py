import re

from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.workbook import Workbook


def resolve_named_chart_reference(
    workbook: Workbook,
    reference: str,
) -> str | None:
    normalized = reference.lstrip("=")
    external_name_match = re.fullmatch(r"\[\d+\]!(.+)", normalized)
    if external_name_match is None:
        return normalized if "[" not in normalized else None

    defined_name = workbook.defined_names.get(external_name_match.group(1))
    expression = getattr(defined_name, "attr_text", None)
    if not isinstance(expression, str):
        return None
    if not expression.upper().startswith("OFFSET("):
        return expression if "!" in expression and "[" not in expression else None

    arguments = split_function_arguments(expression[7:-1])
    if len(arguments) < 3 or "!" not in arguments[0]:
        return None
    sheet_token, cell_token = arguments[0].rsplit("!", 1)
    sheet_name = sheet_token.strip("'").replace("''", "'")
    if sheet_name not in workbook.sheetnames:
        return None

    try:
        min_column, min_row, _, _ = range_boundaries(cell_token.replace("$", ""))
        row_offset = resolve_integer_argument(workbook, arguments[1])
        column_offset = resolve_integer_argument(workbook, arguments[2])
        height = resolve_integer_argument(workbook, arguments[3]) if len(arguments) > 3 else 1
        width = resolve_integer_argument(workbook, arguments[4]) if len(arguments) > 4 else 1
    except (TypeError, ValueError):
        return None

    start_row = min_row + row_offset
    start_column = min_column + column_offset
    if start_row < 1 or start_column < 1 or height < 1 or width < 1:
        return None
    end_row = start_row + height - 1
    end_column = start_column + width - 1
    start_cell = _absolute_cell(start_column, start_row)
    end_cell = _absolute_cell(end_column, end_row)
    escaped_sheet_name = sheet_name.replace("'", "''")
    return f"'{escaped_sheet_name}'!{start_cell}:{end_cell}"


def resolve_integer_argument(workbook: Workbook, argument: str) -> int:
    normalized = argument.strip()
    try:
        return int(float(normalized))
    except ValueError:
        pass
    if "!" not in normalized:
        raise ValueError("OFFSET argument is not numeric or a cell reference")
    sheet_token, cell_token = normalized.rsplit("!", 1)
    sheet_name = sheet_token.strip("'").replace("''", "'")
    value = workbook[sheet_name][cell_token.replace("$", "")].value
    if not isinstance(value, (int, float)):
        raise ValueError("OFFSET argument cell does not contain a number")
    return int(value)


def split_function_arguments(arguments: str) -> list[str]:
    parts: list[str] = []
    start = 0
    depth = 0
    for index, character in enumerate(arguments):
        if character == "(":
            depth += 1
        elif character == ")":
            depth = max(0, depth - 1)
        elif character == "," and depth == 0:
            parts.append(arguments[start:index].strip())
            start = index + 1
    parts.append(arguments[start:].strip())
    return parts


def _absolute_cell(column: int, row: int) -> str:
    return f"${get_column_letter(column)}${row}"
