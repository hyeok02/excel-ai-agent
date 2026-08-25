from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.semantic_models import (
    SemanticClassification,
    SemanticReason,
    SemanticRole,
)
from app.services.workbook_details.cell_values import (
    cell_fill_color,
    is_merged_cell,
    json_value,
    string_or_none,
)
from app.services.workbook_details.models import CellSnapshot


def snapshot_range(
    worksheet: Worksheet,
    value_worksheet: Worksheet | None,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
    region_semantic: SemanticClassification | None = None,
) -> list[list[CellSnapshot]]:
    return [
        [
            _snapshot_cell(
                worksheet.cell(row=row, column=column),
                value_worksheet.cell(row=row, column=column)
                if value_worksheet is not None
                else None,
                worksheet,
                region_semantic,
            )
            for column in range(min_column, max_column + 1)
        ]
        for row in range(min_row, max_row + 1)
    ]


def _snapshot_cell(
    cell: object,
    value_cell: object | None,
    worksheet: Worksheet,
    region_semantic: SemanticClassification | None,
) -> CellSnapshot:
    address = str(getattr(cell, "coordinate"))
    raw_value = getattr(cell, "value", None)
    is_formula = getattr(cell, "data_type", None) == "f" and isinstance(
        raw_value, str
    )
    return CellSnapshot(
        address=address,
        value=None if is_formula else json_value(raw_value),
        formula=raw_value if is_formula else None,
        cached_value=(
            json_value(getattr(value_cell, "value", None)) if is_formula else None
        ),
        number_format=string_or_none(getattr(cell, "number_format", None)),
        bold=bool(getattr(getattr(cell, "font", None), "bold", False)),
        fill_color=cell_fill_color(cell),
        horizontal_alignment=string_or_none(
            getattr(getattr(cell, "alignment", None), "horizontal", None)
        ),
        merged=is_merged_cell(worksheet, address),
        semantic=_cell_semantic(
            address, raw_value, is_formula, region_semantic, worksheet
        ),
    )


def _cell_semantic(
    address: str,
    value: object,
    is_formula: bool,
    region_semantic: SemanticClassification | None,
    worksheet: Worksheet,
) -> SemanticClassification | None:
    if is_formula:
        return SemanticClassification(
            role=SemanticRole.FORMULA,
            confidence=1.0,
            reasons=(
                SemanticReason(
                    code="formula_cell",
                    message="Excel 수식이 입력된 계산 셀",
                    evidence_cells=(address,),
                ),
            ),
        )
    if value is not None or is_merged_cell(worksheet, address):
        return region_semantic
    return None
