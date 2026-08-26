from dataclasses import dataclass

from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
)
from openpyxl.worksheet.worksheet import Worksheet

from app.services.region_detector import CellRegion
from app.services.semantic_models import SemanticRole
from app.services.workbook_details.models import ColumnSchemaSummary, RegionSummary
from app.services.workbook_details.standard_fields import classify_standard_field
from app.services.workbook_details.unit_detection import infer_data_type, infer_unit

DATA_ROLES = {
    SemanticRole.DATA,
    SemanticRole.FORMULA,
    SemanticRole.TOTAL,
    SemanticRole.INPUT,
    SemanticRole.CALCULATION,
    SemanticRole.OUTPUT,
}
MAX_SAMPLE_COUNT = 100


@dataclass(frozen=True)
class _ColumnSource:
    column: int
    source_range: str
    labels: list[str]
    min_row: int
    max_row: int


def build_column_schemas(
    worksheet: Worksheet,
    value_worksheet: Worksheet,
    detected: list[CellRegion],
    summaries: list[RegionSummary],
) -> list[ColumnSchemaSummary]:
    sources = _region_sources(detected, summaries) + _table_sources(worksheet)
    schemas: list[ColumnSchemaSummary] = []
    seen: set[tuple[str, int]] = set()
    for source in sources:
        key = (source.source_range, source.column)
        if key in seen:
            continue
        seen.add(key)
        values, formats = _observations(worksheet, value_worksheet, source)
        data_type = infer_data_type(values, formats)
        field, field_confidence, field_evidence = classify_standard_field(
            source.labels, data_type
        )
        unit_type, unit_label, unit_confidence, unit_evidence = infer_unit(
            source.labels, values, formats, field
        )
        schemas.append(
            ColumnSchemaSummary(
                column=get_column_letter(source.column),
                source_range=source.source_range,
                header_path=source.labels,
                display_name=" > ".join(source.labels),
                standard_field=field,
                data_type=data_type,
                unit_type=unit_type,
                unit_label=unit_label,
                confidence=round((field_confidence + unit_confidence) / 2, 2),
                evidence=field_evidence + unit_evidence,
            )
        )
    return schemas


def _region_sources(
    detected: list[CellRegion], summaries: list[RegionSummary]
) -> list[_ColumnSource]:
    sources: list[_ColumnSource] = []
    for summary in summaries:
        if not summary.header_paths:
            continue
        min_col, min_row, max_col, header_max_row = range_boundaries(
            f"{summary.start_cell}:{summary.end_cell}"
        )
        related = _nearest_data_regions(detected, header_max_row, min_col, max_col)
        max_row = max((_bounds(item)[3] for item in related), default=header_max_row + 20)
        source_range = f"{summary.start_cell}:{_cell(max_col, max_row)}"
        for path in summary.header_paths:
            column = column_index_from_string(path.column)
            sources.append(_ColumnSource(column, source_range, path.labels, header_max_row + 1, max_row))
    return sources


def _nearest_data_regions(
    regions: list[CellRegion], header_row: int, min_col: int, max_col: int
) -> list[CellRegion]:
    candidates = []
    for region in regions:
        left, top, right, _ = _bounds(region)
        role = region.semantic.role if region.semantic else None
        if role in DATA_ROLES and top > header_row and left <= max_col and right >= min_col:
            candidates.append(region)
    if not candidates:
        return []
    nearest_row = min(_bounds(item)[1] for item in candidates)
    return [item for item in candidates if _bounds(item)[1] == nearest_row]


def _table_sources(worksheet: Worksheet) -> list[_ColumnSource]:
    sources = []
    for table in worksheet.tables.values():
        min_col, min_row, _, max_row = range_boundaries(table.ref)
        for offset, column in enumerate(table.tableColumns):
            sources.append(_ColumnSource(min_col + offset, table.ref, [column.name], min_row + 1, max_row))
    return sources


def _observations(
    worksheet: Worksheet, values: Worksheet, source: _ColumnSource
) -> tuple[list[object], list[str]]:
    rows = range(source.min_row, min(source.max_row, worksheet.max_row) + 1)
    value_samples, formats = [], []
    for row in rows:
        raw_cell = worksheet.cell(row, source.column)
        value = values.cell(row, source.column).value
        if value is not None and len(value_samples) < MAX_SAMPLE_COUNT:
            value_samples.append(value)
        if raw_cell.number_format and raw_cell.number_format not in formats:
            formats.append(raw_cell.number_format)
    return value_samples, formats


def _bounds(region: CellRegion) -> tuple[int, int, int, int]:
    return range_boundaries(f"{region.start_cell}:{region.end_cell}")


def _cell(column: int, row: int) -> str:
    return f"{get_column_letter(column)}{row}"
