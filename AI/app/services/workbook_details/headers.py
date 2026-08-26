from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.semantic_models import SemanticRole
from app.services.workbook_details.header_labels import header_label
from app.services.workbook_details.merged_headers import resolved_header_rows
from app.services.workbook_details.models import HeaderPathSummary

HEADER_SCAN_ROWS = 4
HEADER_SCAN_COLUMNS = 12


def region_title(
    worksheet: Worksheet,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
) -> str | None:
    for row in worksheet.iter_rows(
        min_row=min_row,
        max_row=min(max_row, min_row + 3),
        min_col=min_column,
        max_col=max_column,
    ):
        for cell in row:
            label = header_label(cell.value)
            if label is not None:
                return label
    return None


def header_paths(
    worksheet: Worksheet,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
    semantic_role: SemanticRole | None,
) -> list[HeaderPathSummary]:
    if semantic_role is not SemanticRole.HEADER:
        return []
    header_max_row = min(max_row, min_row + HEADER_SCAN_ROWS - 1)
    column_max = min(max_column, min_column + HEADER_SCAN_COLUMNS - 1)
    propagated_rows = resolved_header_rows(
        worksheet,
        min_row,
        header_max_row,
        min_column,
        column_max,
    )
    paths = []
    for index, column_number in enumerate(range(min_column, column_max + 1)):
        labels = []
        for row_labels in propagated_rows:
            label = row_labels[index]
            if label is not None and (not labels or labels[-1] != label):
                labels.append(label)
        if labels:
            paths.append(
                HeaderPathSummary(column=get_column_letter(column_number), labels=labels)
            )
    return paths
