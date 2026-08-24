from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
import re
from typing import TypeAlias

from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

from app.services.region_detector import CellRegion
from app.services.analysis_inclusion import (
    AnalysisInclusion,
    INCLUDED_POPULATED_REGION,
)
from app.services.semantic_models import SemanticClassification

CellValue: TypeAlias = str | int | float | bool | None

REGION_PREVIEW_ROWS = 8
REGION_PREVIEW_COLUMNS = 8
TABLE_PREVIEW_ROWS = 8
TABLE_PREVIEW_COLUMNS = 12
CHART_SERIES_LIMIT = 12
CHART_SAMPLE_VALUES = 12


@dataclass(frozen=True)
class CellSnapshot:
    address: str
    value: CellValue
    formula: str | None
    cached_value: CellValue = None
    number_format: str | None = None
    bold: bool = False
    fill_color: str | None = None
    horizontal_alignment: str | None = None
    merged: bool = False
    semantic: SemanticClassification | None = None


@dataclass(frozen=True)
class HeaderPathSummary:
    column: str
    labels: list[str]


@dataclass(frozen=True)
class RegionSummary:
    start_cell: str
    end_cell: str
    cell_count: int
    title: str | None
    row_count: int
    column_count: int
    merged_ranges: list[str]
    header_paths: list[HeaderPathSummary]
    preview_rows: list[list[CellSnapshot]]
    is_truncated: bool
    analysis_inclusion: AnalysisInclusion = INCLUDED_POPULATED_REGION
    semantic: SemanticClassification | None = None


@dataclass(frozen=True)
class TableSummary:
    name: str
    display_name: str
    reference: str
    headers: list[str]
    row_count: int
    column_count: int
    preview_rows: list[list[CellSnapshot]]
    is_truncated: bool


@dataclass(frozen=True)
class ChartSeriesSummary:
    title: str | None
    categories_reference: str | None
    values_reference: str | None
    category_samples: list[CellValue]
    value_samples: list[CellValue]


@dataclass(frozen=True)
class ChartSummary:
    title: str | None
    chart_type: str
    anchor_cell: str | None
    series_count: int
    series: list[ChartSeriesSummary]
    is_truncated: bool


def summarize_regions(
    worksheet: Worksheet,
    regions: list[CellRegion],
    value_worksheet: Worksheet | None = None,
) -> list[RegionSummary]:
    summaries: list[RegionSummary] = []

    for region in regions:
        min_column, min_row, max_column, max_row = range_boundaries(
            f"{region.start_cell}:{region.end_cell}"
        )
        preview_max_row = min(max_row, min_row + REGION_PREVIEW_ROWS - 1)
        preview_max_column = min(
            max_column,
            min_column + REGION_PREVIEW_COLUMNS - 1,
        )
        summaries.append(
            RegionSummary(
                start_cell=region.start_cell,
                end_cell=region.end_cell,
                cell_count=region.cell_count,
                title=_region_title(worksheet, min_row, max_row, min_column, max_column),
                row_count=max_row - min_row + 1,
                column_count=max_column - min_column + 1,
                merged_ranges=_intersecting_merged_ranges(
                    worksheet,
                    min_row,
                    max_row,
                    min_column,
                    max_column,
                ),
                header_paths=_header_paths(
                    worksheet,
                    min_row,
                    max_row,
                    min_column,
                    max_column,
                ),
                preview_rows=_snapshot_range(
                    worksheet,
                    value_worksheet,
                    min_row,
                    preview_max_row,
                    min_column,
                    preview_max_column,
                ),
                is_truncated=(
                    preview_max_row < max_row or preview_max_column < max_column
                ),
            )
        )

    return summaries


def summarize_tables(
    worksheet: Worksheet,
    value_worksheet: Worksheet | None = None,
) -> list[TableSummary]:
    summaries: list[TableSummary] = []

    for table in worksheet.tables.values():
        min_column, min_row, max_column, max_row = range_boundaries(table.ref)
        preview_max_row = min(max_row, min_row + TABLE_PREVIEW_ROWS - 1)
        preview_max_column = min(
            max_column,
            min_column + TABLE_PREVIEW_COLUMNS - 1,
        )
        summaries.append(
            TableSummary(
                name=table.name,
                display_name=table.displayName,
                reference=table.ref,
                headers=[column.name for column in table.tableColumns],
                row_count=max_row - min_row + 1,
                column_count=max_column - min_column + 1,
                preview_rows=_snapshot_range(
                    worksheet,
                    value_worksheet,
                    min_row,
                    preview_max_row,
                    min_column,
                    preview_max_column,
                ),
                is_truncated=(
                    preview_max_row < max_row or preview_max_column < max_column
                ),
            )
        )

    return summaries


def summarize_charts(
    workbook: Workbook,
    worksheet: Worksheet,
    value_workbook: Workbook | None = None,
) -> list[ChartSummary]:
    summaries: list[ChartSummary] = []

    for chart in worksheet._charts:
        all_series = list(chart.ser)
        series_summaries = [
            _summarize_chart_series(value_workbook or workbook, series)
            for series in all_series[:CHART_SERIES_LIMIT]
        ]
        summaries.append(
            ChartSummary(
                title=_chart_title(chart),
                chart_type=type(chart).__name__,
                anchor_cell=_chart_anchor(chart),
                series_count=len(all_series),
                series=series_summaries,
                is_truncated=len(all_series) > CHART_SERIES_LIMIT,
            )
        )

    return summaries


def _snapshot_range(
    worksheet: Worksheet,
    value_worksheet: Worksheet | None,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
) -> list[list[CellSnapshot]]:
    return [
        [
            _snapshot_cell(
                worksheet.cell(row=row, column=column),
                value_worksheet.cell(row=row, column=column)
                if value_worksheet is not None
                else None,
                worksheet,
            )
            for column in range(min_column, max_column + 1)
        ]
        for row in range(min_row, max_row + 1)
    ]


def _snapshot_cell(
    cell: object,
    value_cell: object | None,
    worksheet: Worksheet,
) -> CellSnapshot:
    address = str(getattr(cell, "coordinate"))
    raw_value = getattr(cell, "value", None)
    is_formula = getattr(cell, "data_type", None) == "f" and isinstance(
        raw_value, str
    )
    return CellSnapshot(
        address=address,
        value=None if is_formula else _json_value(raw_value),
        formula=raw_value if is_formula else None,
        cached_value=(
            _json_value(getattr(value_cell, "value", None)) if is_formula else None
        ),
        number_format=_string_or_none(getattr(cell, "number_format", None)),
        bold=bool(getattr(getattr(cell, "font", None), "bold", False)),
        fill_color=_cell_fill_color(cell),
        horizontal_alignment=_string_or_none(
            getattr(getattr(cell, "alignment", None), "horizontal", None)
        ),
        merged=_is_merged_cell(worksheet, address),
    )


def _region_title(
    worksheet: Worksheet,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
) -> str | None:
    for row in worksheet.iter_rows(
        min_row=min_row,
        max_row=min(max_row, min_row + 3),
        min_col=min_column,
        max_col=max_column,
    ):
        for cell in row:
            label = _header_label(cell.value)
            if label is not None:
                return label
    return None


def _header_paths(
    worksheet: Worksheet,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
) -> list[HeaderPathSummary]:
    header_max_row = min(max_row, min_row + 3)
    column_max = min(max_column, min_column + TABLE_PREVIEW_COLUMNS - 1)
    propagated_rows: list[list[str | None]] = []

    for row_number in range(min_row, header_max_row + 1):
        row_labels: list[str | None] = []
        previous_label: str | None = None
        text_count = 0
        data_count = 0
        for column_number in range(min_column, column_max + 1):
            cell = worksheet.cell(row=row_number, column=column_number)
            label = _header_label(cell.value)
            if label is not None:
                text_count += 1
            elif cell.value is not None:
                data_count += 1
            if label is None and _is_merged_cell(worksheet, cell.coordinate):
                label = previous_label
            if label is not None:
                previous_label = label
            row_labels.append(label)
        if propagated_rows and (text_count == 0 or data_count > text_count):
            break
        propagated_rows.append(row_labels)

    paths: list[HeaderPathSummary] = []
    for index, column_number in enumerate(range(min_column, column_max + 1)):
        labels: list[str] = []
        for row_labels in propagated_rows:
            label = row_labels[index]
            if label is not None and (not labels or labels[-1] != label):
                labels.append(label)
        if labels:
            paths.append(
                HeaderPathSummary(
                    column=get_column_letter(column_number),
                    labels=labels,
                )
            )
    return paths


def _header_label(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    normalized = value.strip()
    if not normalized or normalized.startswith("=") or len(normalized) > 120:
        return None
    return normalized


def _intersecting_merged_ranges(
    worksheet: Worksheet,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
) -> list[str]:
    matches: list[str] = []
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.max_row >= min_row
            and merged_range.min_row <= max_row
            and merged_range.max_col >= min_column
            and merged_range.min_col <= max_column
        ):
            matches.append(str(merged_range))
            if len(matches) == 20:
                break
    return matches


def _is_merged_cell(worksheet: Worksheet, coordinate: str) -> bool:
    return any(coordinate in merged_range for merged_range in worksheet.merged_cells.ranges)


def _cell_fill_color(cell: object) -> str | None:
    fill = getattr(cell, "fill", None)
    if getattr(fill, "fill_type", None) is None:
        return None
    color = getattr(fill, "fgColor", None)
    color_type = getattr(color, "type", None)
    value = getattr(color, color_type, None) if color_type else None
    return str(value) if value not in (None, "00000000", "000000") else None


def _json_value(value: object) -> CellValue:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return str(value)


def _summarize_chart_series(
    workbook: Workbook,
    series: object,
) -> ChartSeriesSummary:
    title_reference = _nested_attr(series, "tx", "strRef", "f")
    title_value = _nested_attr(series, "tx", "v")
    categories_reference = (
        _nested_attr(series, "cat", "strRef", "f")
        or _nested_attr(series, "cat", "numRef", "f")
        or _nested_attr(series, "cat", "multiLvlStrRef", "f")
    )
    values_reference = _nested_attr(series, "val", "numRef", "f")
    resolved_title = _resolve_reference_values(workbook, title_reference, 1)

    return ChartSeriesSummary(
        title=(
            str(resolved_title[0])
            if resolved_title and resolved_title[0] is not None
            else str(title_value) if title_value is not None else None
        ),
        categories_reference=_string_or_none(categories_reference),
        values_reference=_string_or_none(values_reference),
        category_samples=_resolve_reference_values(
            workbook,
            categories_reference,
            CHART_SAMPLE_VALUES,
        ),
        value_samples=_resolve_reference_values(
            workbook,
            values_reference,
            CHART_SAMPLE_VALUES,
        ),
    )


def _resolve_reference_values(
    workbook: Workbook,
    reference: object,
    limit: int,
) -> list[CellValue]:
    if not isinstance(reference, str):
        return []

    normalized_reference = _resolve_named_chart_reference(workbook, reference)
    if normalized_reference is None or "!" not in normalized_reference:
        return []

    sheet_token, range_token = normalized_reference.lstrip("=").rsplit("!", 1)
    sheet_name = sheet_token.strip("'").replace("''", "'")
    if sheet_name not in workbook.sheetnames:
        return []

    try:
        min_column, min_row, max_column, max_row = range_boundaries(
            range_token.replace("$", "")
        )
    except ValueError:
        return []

    values: list[CellValue] = []
    worksheet = workbook[sheet_name]
    for row in worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_column,
        max_col=max_column,
    ):
        for cell in row:
            values.append(_json_value(cell.value))
            if len(values) == limit:
                return values
    return values


def _resolve_named_chart_reference(
    workbook: Workbook,
    reference: str,
) -> str | None:
    normalized = reference.lstrip("=")
    external_name_match = re.fullmatch(r"\[\d+\]!(.+)", normalized)
    if external_name_match is None:
        return normalized if "[" not in normalized else None

    defined_name = workbook.defined_names.get(external_name_match.group(1))
    expression = getattr(defined_name, "attr_text", None)
    if not isinstance(expression, str):
        return None

    if not expression.upper().startswith("OFFSET("):
        return expression if "!" in expression and "[" not in expression else None

    offset_arguments = _split_function_arguments(expression[7:-1])
    if len(offset_arguments) < 3:
        return None

    base_reference = offset_arguments[0]
    if "!" not in base_reference:
        return None
    sheet_token, cell_token = base_reference.rsplit("!", 1)
    sheet_name = sheet_token.strip("'").replace("''", "'")
    if sheet_name not in workbook.sheetnames:
        return None

    try:
        min_column, min_row, _, _ = range_boundaries(cell_token.replace("$", ""))
        row_offset = _resolve_integer_argument(workbook, offset_arguments[1])
        column_offset = _resolve_integer_argument(workbook, offset_arguments[2])
        height = (
            _resolve_integer_argument(workbook, offset_arguments[3])
            if len(offset_arguments) > 3
            else 1
        )
        width = (
            _resolve_integer_argument(workbook, offset_arguments[4])
            if len(offset_arguments) > 4
            else 1
        )
    except (TypeError, ValueError):
        return None

    start_row = min_row + row_offset
    start_column = min_column + column_offset
    if start_row < 1 or start_column < 1 or height < 1 or width < 1:
        return None
    end_row = start_row + height - 1
    end_column = start_column + width - 1
    start_cell = f"{get_column_letter(start_column)}{start_row}"
    end_cell = f"{get_column_letter(end_column)}{end_row}"
    escaped_sheet_name = sheet_name.replace("'", "''")
    return f"'{escaped_sheet_name}'!${start_cell.replace(str(start_row), '$' + str(start_row))}:${end_cell.replace(str(end_row), '$' + str(end_row))}"


def _resolve_integer_argument(workbook: Workbook, argument: str) -> int:
    normalized = argument.strip()
    try:
        return int(float(normalized))
    except ValueError:
        pass

    if "!" not in normalized:
        raise ValueError("OFFSET argument is not numeric or a cell reference")
    sheet_token, cell_token = normalized.rsplit("!", 1)
    sheet_name = sheet_token.strip("'").replace("''", "'")
    value = workbook[sheet_name][cell_token.replace("$", "")].value
    if not isinstance(value, (int, float)):
        raise ValueError("OFFSET argument cell does not contain a number")
    return int(value)


def _split_function_arguments(arguments: str) -> list[str]:
    parts: list[str] = []
    start = 0
    depth = 0
    for index, character in enumerate(arguments):
        if character == "(":
            depth += 1
        elif character == ")":
            depth = max(0, depth - 1)
        elif character == "," and depth == 0:
            parts.append(arguments[start:index].strip())
            start = index + 1
    parts.append(arguments[start:].strip())
    return parts


def _chart_title(chart: object) -> str | None:
    title = getattr(chart, "title", None)
    if isinstance(title, str):
        return title

    rich_text = _nested_attr(title, "tx", "rich")
    paragraphs = getattr(rich_text, "p", []) if rich_text is not None else []
    parts: list[str] = []
    for paragraph in paragraphs:
        for run in getattr(paragraph, "r", []) or []:
            text = getattr(run, "t", None)
            if text:
                parts.append(str(text))
        for field in getattr(paragraph, "fld", []) or []:
            text = getattr(field, "t", None)
            if text:
                parts.append(str(text))
    return "".join(parts) or None


def _chart_anchor(chart: object) -> str | None:
    marker = getattr(getattr(chart, "anchor", None), "_from", None)
    row = getattr(marker, "row", None)
    column = getattr(marker, "col", None)
    if not isinstance(row, int) or not isinstance(column, int):
        return None
    return f"{get_column_letter(column + 1)}{row + 1}"


def _nested_attr(value: object, *attributes: str) -> object | None:
    current = value
    for attribute in attributes:
        if current is None:
            return None
        current = getattr(current, attribute, None)
    return current


def _string_or_none(value: object) -> str | None:
    return str(value) if value is not None else None
