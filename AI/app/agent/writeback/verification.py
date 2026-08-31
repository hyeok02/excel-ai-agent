import hashlib
from io import BytesIO
from zipfile import ZipFile

from openpyxl import load_workbook

from app.agent.writeback.models import VerificationCheck


def workbook_fingerprint(content: bytes, keep_vba: bool) -> dict[str, object]:
    workbook = load_workbook(BytesIO(content), data_only=False, keep_vba=keep_vba)
    try:
        return {
            "sheets": tuple(workbook.sheetnames),
            "formulas": _formulas(workbook),
            "merged": _merged(workbook),
            "styles": _styles(workbook),
        }
    finally:
        workbook.close()


def compare_fingerprints(before, after, changes) -> list[VerificationCheck]:
    targets = {(change.sheet_name, change.reference) for change in changes}
    unchanged_styles = all(
        before["styles"].get(target) == after["styles"].get(target) for target in targets
    )
    return [
        _check("sheet_structure", before["sheets"] == after["sheets"], "시트 구성 보존"),
        _check("formulas", before["formulas"] == after["formulas"], "전체 수식 보존"),
        _check("merged_cells", before["merged"] == after["merged"], "병합 영역 보존"),
        _check("styles", unchanged_styles, "변경 셀 서식 보존"),
    ]


def package_checks(
    before: bytes, after: bytes, changed_paths: set[str]
) -> list[VerificationCheck]:
    with ZipFile(BytesIO(before)) as original, ZipFile(BytesIO(after)) as modified:
        original_names, modified_names = set(original.namelist()), set(modified.namelist())
        unchanged = original_names - changed_paths
        parts_preserved = all(
            original.read(name) == modified.read(name) for name in unchanged
        )
        features_preserved = all(
            _features(original.read(path)) == _features(modified.read(path))
            for path in changed_paths
        )
    return [
        _check("package_structure", original_names == modified_names, "파일 구성 요소 보존"),
        _check("unchanged_parts", parts_preserved, "변경 시트 외 원본 부품 보존"),
        _check("excel_extensions", features_preserved, "유효성·확장 기능 보존"),
    ]


def vba_digest(content: bytes) -> str | None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            name = next((item for item in archive.namelist() if item.endswith("vbaProject.bin")), None)
            return hashlib.sha256(archive.read(name)).hexdigest() if name else None
    except Exception:
        return None


def add_macro_check(checks, before: bytes, after: bytes, keep_vba: bool) -> None:
    if not keep_vba:
        checks.append(_check("macros", True, "매크로가 없는 .xlsx 파일"))
        return
    original, modified = vba_digest(before), vba_digest(after)
    checks.append(_check("macros", original == modified, "VBA 프로젝트 보존"))


def _formulas(workbook) -> dict[tuple[str, str], str]:
    return {
        (sheet.title, cell.coordinate): cell.value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }


def _merged(workbook) -> dict[str, tuple[str, ...]]:
    return {
        sheet.title: tuple(sorted(str(item) for item in sheet.merged_cells.ranges))
        for sheet in workbook.worksheets
    }


def _styles(workbook) -> dict[tuple[str, str], int]:
    return {
        (sheet.title, cell.coordinate): cell.style_id
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    }


def _check(name: str, passed: bool, detail: str) -> VerificationCheck:
    return VerificationCheck(name=name, passed=passed, detail=detail)


def _features(content: bytes) -> tuple[bytes, ...]:
    return tuple(
        match.group(0)
        for tag in (b"dataValidations", b"conditionalFormatting", b"extLst")
        for match in __import__("re").finditer(
            rb"<" + tag + rb"\b.*?</" + tag + rb">", content, __import__("re").DOTALL
        )
    )
