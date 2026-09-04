from pathlib import Path
import re

from openpyxl.cell.cell import MergedCell
from app.services.workbook_loading import close_workbook, load_workbook_for_reading

from app.agent.writeback.models import VerificationCheck, WritebackChange, WritebackManifest
from app.agent.writeback.package_editor import changed_worksheet_paths, patch_workbook_package
from app.agent.writeback.verification import (
    add_macro_check,
    compare_fingerprints,
    package_checks,
    workbook_fingerprint,
)


class UnsafeWritebackError(ValueError):
    """Raised when a requested workbook change cannot be applied safely."""


def apply_writeback(
    filename: str, content: bytes, changes: list[WritebackChange]
) -> tuple[bytes, WritebackManifest]:
    if not changes or len(changes) > 50:
        raise UnsafeWritebackError("한 번에 1~50개의 검증된 변경만 적용할 수 있습니다.")
    keep_vba = Path(filename).suffix.lower() == ".xlsm"
    before = workbook_fingerprint(content, keep_vba)
    workbook = load_workbook_for_reading(content, keep_vba=keep_vba)
    try:
        _validate_changes(workbook, changes)
    finally:
        close_workbook(workbook)
    changed_paths = changed_worksheet_paths(content, changes)
    try:
        modified = patch_workbook_package(content, changes)
    except ValueError as exception:
        raise UnsafeWritebackError(str(exception)) from exception
    after = workbook_fingerprint(modified, keep_vba)
    checks = compare_fingerprints(before, after, changes)
    checks.extend(package_checks(content, modified, changed_paths))
    add_macro_check(checks, content, modified, keep_vba)
    checks.append(_verify_values(modified, changes, keep_vba))
    manifest = WritebackManifest(
        changed_cells=[f"{item.sheet_name}!{item.reference}" for item in changes],
        checks=checks,
        verified=all(check.passed for check in checks),
    )
    if not manifest.verified:
        raise UnsafeWritebackError("변경 후 보존 검증에 실패해 수정본을 생성하지 않았습니다.")
    return modified, manifest


def _validate_changes(workbook, changes: list[WritebackChange]) -> None:
    seen = set()
    for change in changes:
        key = (change.sheet_name, change.reference.upper())
        if key in seen or change.sheet_name not in workbook.sheetnames:
            raise UnsafeWritebackError("중복되거나 존재하지 않는 변경 대상입니다.")
        seen.add(key)
        cell = workbook[change.sheet_name][change.reference]
        if isinstance(cell, MergedCell):
            raise UnsafeWritebackError("병합 영역의 시작 셀만 수정할 수 있습니다.")
        old_formula = isinstance(cell.value, str) and cell.value.startswith("=")
        new_formula = _formula(change.new_value)
        if old_formula and new_formula is None:
            raise UnsafeWritebackError("기존 수식은 승인된 새 수식으로만 변경할 수 있습니다.")
        if _comparable(cell.value, change.value_type) != _comparable(
            change.old_value, change.value_type
        ):
            raise UnsafeWritebackError("승인한 기존 값과 현재 원본 값이 다릅니다.")
        if new_formula and _unsafe_formula(new_formula):
            raise UnsafeWritebackError("외부 연결이나 실행 기능이 포함된 수식은 허용하지 않습니다.")
        if new_formula and len(new_formula) > 8192:
            raise UnsafeWritebackError("Excel 수식 최대 길이를 초과했습니다.")
        if isinstance(change.new_value, str) and len(change.new_value) > 32_767:
            raise UnsafeWritebackError("Excel 셀의 최대 문자열 길이를 초과했습니다.")


def _verify_values(content, changes, keep_vba) -> VerificationCheck:
    workbook = load_workbook_for_reading(content, keep_vba=keep_vba)
    try:
        passed = all(
            _comparable(
                workbook[item.sheet_name][item.reference].value, item.value_type
            )
            == _comparable(item.new_value, item.value_type)
            for item in changes
        )
    finally:
        close_workbook(workbook)
    return VerificationCheck(name="changed_values", passed=passed, detail="승인한 값 반영")


def _formula(value: object) -> str | None:
    return value.strip() if isinstance(value, str) and value.lstrip().startswith("=") else None


def _unsafe_formula(formula: str) -> bool:
    return bool(
        re.search(
            r"(?:\[|https?://|\\\\|\||\b(?:WEBSERVICE|HYPERLINK|RTD|CALL|REGISTER\.ID|EXEC)\s*\()",
            formula,
            re.IGNORECASE,
        )
    )


def _comparable(value: object, value_type: str | None = None) -> object:
    if value_type == "date" and hasattr(value, "date"):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value
