from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook

from app.agent.query.cell_values import formula_text, safe_indexed_value
from app.services.provenance import AnalysisEvidence, EvidenceKind

MAX_INDEXED_CELLS = 50_000
MAX_ROWS_PER_SHEET = 10_000


@dataclass(frozen=True)
class IndexedCell:
    sheet_name: str
    address: str
    value: str | int | float | bool | None
    formula: str | None
    value_type: str = "text"

    @property
    def reference(self) -> str:
        return f"{self.sheet_name}!{self.address}"

    def evidence(self) -> AnalysisEvidence:
        return AnalysisEvidence(
            kind=EvidenceKind.FORMULA if self.formula else EvidenceKind.CELL,
            sheet_name=self.sheet_name,
            reference=self.address,
            description="질문과 관련해 원본 Excel에서 조회한 셀",
            value=self.value,
            formula=self.formula,
        )


@dataclass(frozen=True)
class IndexedRow:
    sheet_name: str
    row_number: int
    cells: tuple[IndexedCell, ...]

    @property
    def search_text(self) -> str:
        return " ".join(str(cell.value or cell.formula or "") for cell in self.cells).casefold()


@dataclass(frozen=True)
class WorkbookDataIndex:
    filename: str
    rows: tuple[IndexedRow, ...]
    indexed_cell_count: int
    truncated: bool


def build_workbook_data_index(
    filename: str, content: bytes, included_sheets: set[str] | None = None
) -> WorkbookDataIndex:
    formulas = load_workbook(BytesIO(content), data_only=False, read_only=True)
    values = load_workbook(BytesIO(content), data_only=True, read_only=True)
    rows: list[IndexedRow] = []
    cell_count = 0
    truncated = False
    try:
        for formula_sheet in formulas.worksheets:
            if included_sheets is not None and formula_sheet.title not in included_sheets:
                continue
            value_sheet = values[formula_sheet.title]
            for row_number, (formula_row, value_row) in enumerate(
                zip(formula_sheet.iter_rows(), value_sheet.iter_rows()), start=1
            ):
                if row_number > MAX_ROWS_PER_SHEET or cell_count >= MAX_INDEXED_CELLS:
                    truncated = True
                    break
                cells = _indexed_cells(formula_sheet.title, formula_row, value_row)
                if cells:
                    remaining = MAX_INDEXED_CELLS - cell_count
                    cells = cells[:remaining]
                    rows.append(IndexedRow(formula_sheet.title, row_number, tuple(cells)))
                    cell_count += len(cells)
            if cell_count >= MAX_INDEXED_CELLS:
                truncated = True
                break
    finally:
        formulas.close()
        values.close()
    return WorkbookDataIndex(filename, tuple(rows), cell_count, truncated)


def _indexed_cells(sheet_name: str, formula_row: tuple, value_row: tuple) -> list[IndexedCell]:
    cells = []
    for formula_cell, value_cell in zip(formula_row, value_row):
        raw = formula_cell.value
        cached = value_cell.value
        if raw is None and cached is None:
            continue
        formula = formula_text(raw)
        value = cached if formula else raw
        safe_value, value_type = safe_indexed_value(value, formula)
        cells.append(
            IndexedCell(
                sheet_name,
                formula_cell.coordinate,
                safe_value,
                formula,
                value_type,
            )
        )
    return cells
