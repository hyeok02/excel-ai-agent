from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from app.services.formula_analyzer import analyze_formulas
from app.services.sheet_classifier import (
    SheetImportance,
    SheetRole,
    classify_sheets,
)
from app.services.worksheet_filter import evaluate_worksheet_inclusion


def _classification_workbook() -> Workbook:
    workbook = Workbook()

    input_sheet = workbook.active
    input_sheet.title = "입력 데이터"
    input_sheet.append(["상품", "수량", "단가"])
    input_sheet.append(["노트북", 2, 1_500_000])
    input_sheet.append(["모니터", 3, 350_000])

    calculation_sheet = workbook.create_sheet("중간 계산")
    calculation_sheet.append(["상품", "금액"])
    calculation_sheet["A2"] = "노트북"
    calculation_sheet["B2"] = "='입력 데이터'!B2*'입력 데이터'!C2"
    calculation_sheet["A3"] = "모니터"
    calculation_sheet["B3"] = "='입력 데이터'!B3*'입력 데이터'!C3"

    output_sheet = workbook.create_sheet("경영 요약")
    output_sheet.append(["항목", "결과"])
    output_sheet["A2"] = "총 금액"
    output_sheet["B2"] = "=SUM('중간 계산'!B2:B3)"
    chart = BarChart()
    chart.add_data(Reference(output_sheet, min_col=2, min_row=1, max_row=2))
    output_sheet.add_chart(chart, "D2")

    guide_sheet = workbook.create_sheet("사용 안내")
    guide_sheet["A1"] = "워크북 사용 안내"
    guide_sheet["A2"] = "입력 데이터 시트의 수량과 단가를 먼저 수정하세요."
    guide_sheet["A3"] = "중간 계산 시트의 수식은 변경하지 마세요."

    system_sheet = workbook.create_sheet("__snlofficequeries")
    system_sheet["A1"] = "애드인 캐시 데이터"

    return workbook


def test_classifies_input_calculation_output_documentation_and_system_sheets() -> None:
    workbook = _classification_workbook()
    try:
        formulas_by_sheet = {
            worksheet.title: analyze_formulas(worksheet)
            for worksheet in workbook.worksheets
        }
        inclusions_by_sheet = {
            worksheet.title: evaluate_worksheet_inclusion(worksheet)
            for worksheet in workbook.worksheets
        }

        classifications = classify_sheets(
            workbook,
            formulas_by_sheet,
            inclusions_by_sheet,
        )

        assert classifications["입력 데이터"].role is SheetRole.INPUT
        assert classifications["중간 계산"].role is SheetRole.CALCULATION
        assert classifications["경영 요약"].role is SheetRole.OUTPUT
        assert classifications["사용 안내"].role is SheetRole.DOCUMENTATION
        assert classifications["__snlofficequeries"].role is SheetRole.SYSTEM
    finally:
        workbook.close()


def test_returns_explainable_confidence_and_importance() -> None:
    workbook = _classification_workbook()
    try:
        formulas_by_sheet = {
            worksheet.title: analyze_formulas(worksheet)
            for worksheet in workbook.worksheets
        }
        classifications = classify_sheets(
            workbook,
            formulas_by_sheet,
            {
                worksheet.title: evaluate_worksheet_inclusion(worksheet)
                for worksheet in workbook.worksheets
            },
        )

        input_classification = classifications["입력 데이터"]
        calculation_classification = classifications["중간 계산"]
        output_classification = classifications["경영 요약"]
        system_classification = classifications["__snlofficequeries"]

        assert 0.5 <= input_classification.confidence <= 1
        assert input_classification.reasons
        assert any(
            reason.code == "upstream_source"
            for reason in input_classification.reasons
        )
        assert calculation_classification.importance in {
            SheetImportance.HIGH,
            SheetImportance.CRITICAL,
        }
        assert output_classification.importance_score > system_classification.importance_score
        assert system_classification.importance is SheetImportance.LOW
        assert system_classification.reasons[0].code == "system_policy"
    finally:
        workbook.close()
