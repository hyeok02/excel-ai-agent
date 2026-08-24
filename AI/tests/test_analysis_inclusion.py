import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services.analysis_inclusion import AnalysisDecision, AnalysisInclusion
from app.services.worksheet_filter import evaluate_worksheet_inclusion


CONTRACT_PATH = (
    Path(__file__).parents[2] / "contracts" / "analysis-inclusion.schema.json"
)


def test_analysis_decisions_match_shared_contract() -> None:
    contract = json.loads(CONTRACT_PATH.read_text(encoding="utf-8"))

    assert {decision.value for decision in AnalysisDecision} == set(
        contract["$defs"]["analysisDecision"]["enum"]
    )


@pytest.mark.parametrize("field", ["reason_code", "reason"])
def test_rejects_blank_policy_reason_fields(field: str) -> None:
    values = {"reason_code": "business_worksheet", "reason": "분석에 포함"}
    values[field] = "  "

    with pytest.raises(ValueError):
        AnalysisInclusion(decision=AnalysisDecision.INCLUDE, **values)


def test_classifies_visible_hidden_and_cache_worksheets() -> None:
    workbook = Workbook()
    visible = workbook.active
    visible.title = "업무 데이터"
    hidden = workbook.create_sheet("숨김 계산")
    hidden.sheet_state = "hidden"
    addin = workbook.create_sheet("__snlofficequeries")
    cache = workbook.create_sheet("CIOHiddenCacheSheet")

    try:
        assert evaluate_worksheet_inclusion(visible).reason_code == "business_worksheet"
        assert evaluate_worksheet_inclusion(hidden).reason_code == "hidden_worksheet"
        assert evaluate_worksheet_inclusion(addin).reason_code == "addin_cache_worksheet"
        assert evaluate_worksheet_inclusion(cache).reason_code == "system_cache_worksheet"
    finally:
        workbook.close()
