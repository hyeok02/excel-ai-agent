import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from app.services.region_detector import detect_regions


FIXTURE_DIRECTORY = Path(__file__).parent / "fixtures" / "semantic"


def test_detects_regions_separated_by_blank_column() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "왼쪽"
    worksheet["A2"] = 10
    worksheet["D1"] = "오른쪽"
    worksheet["E1"] = 20

    regions = detect_regions(worksheet)

    assert [
        (region.start_cell, region.end_cell, region.cell_count)
        for region in regions
    ] == [
        ("A1", "A2", 2),
        ("D1", "E1", 2),
    ]
    assert [region.semantic.role.value for region in regions] == ["data", "data"]
    workbook.close()


def test_does_not_connect_diagonal_cells() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "첫 번째"
    worksheet["B2"] = "두 번째"

    regions = detect_regions(worksheet)

    assert [
        (region.start_cell, region.end_cell, region.cell_count)
        for region in regions
    ] == [
        ("A1", "A1", 1),
        ("B2", "B2", 1),
    ]
    workbook.close()


def test_ignores_empty_and_whitespace_cells() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = ""
    worksheet["B2"] = "   "

    assert detect_regions(worksheet) == []
    workbook.close()


def test_keeps_boundary_evidence_unique_for_single_cell_after_blank_row() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "첫 번째 영역"
    worksheet["A3"] = "비고: 두 번째 영역"

    regions = detect_regions(worksheet)

    assert [(region.start_cell, region.end_cell) for region in regions] == [
        ("A1", "A1"),
        ("A3", "A3"),
    ]
    blank_row_reason = next(
        reason
        for reason in regions[1].semantic.reasons
        if reason.code == "blank_row_boundary"
    )
    assert blank_row_reason.evidence_cells == ("A3",)
    workbook.close()


def test_detects_title_description_and_note_regions() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells("A1:C1")
    worksheet["A1"] = "프로젝트 손익 현황"
    worksheet["A1"].font = Font(bold=True)
    worksheet.merge_cells("A3:C3")
    worksheet["A3"] = "이 문서는 신규 프로젝트의 예상 손익을 설명합니다."
    worksheet.merge_cells("A5:C5")
    worksheet["A5"] = "비고: 확정 전 수치는 재무팀 검토가 필요합니다."

    regions = detect_regions(worksheet)

    assert [
        (region.start_cell, region.end_cell, region.semantic.role.value)
        for region in regions
    ] == [
        ("A1", "C1", "title"),
        ("A3", "C3", "description"),
        ("A5", "C5", "note"),
    ]
    workbook.close()


def test_uses_fill_style_to_recognize_a_merged_title() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells("A1:C1")
    worksheet["A1"] = "2026년 프로젝트 현황"
    worksheet["A1"].fill = PatternFill("solid", fgColor="DCE8FF")

    regions = detect_regions(worksheet)

    assert len(regions) == 1
    assert regions[0].start_cell == "A1"
    assert regions[0].end_cell == "C1"
    assert regions[0].semantic.role.value == "title"
    assert regions[0].semantic.reasons[0].code == "title_style"
    workbook.close()


def test_detects_expected_semantic_boundaries_from_regression_fixtures() -> None:
    for expectation_path in sorted(FIXTURE_DIRECTORY.glob("*.expected.json")):
        expectation = json.loads(expectation_path.read_text(encoding="utf-8"))
        workbook = load_workbook(FIXTURE_DIRECTORY / expectation["workbook"])
        try:
            for expected_sheet in expectation["sheets"]:
                regions = detect_regions(
                    workbook[expected_sheet["name"]],
                    sheet_role=expected_sheet["sheet_role"],
                )
                actual = {
                    (
                        f"{region.start_cell}:{region.end_cell}",
                        region.semantic.role.value,
                    )
                    for region in regions
                }
                expected = {
                    (region["range"], region["role"])
                    for region in expected_sheet["regions"]
                }

                assert actual == expected
                assert all(region.semantic.reasons for region in regions)
        finally:
            workbook.close()
