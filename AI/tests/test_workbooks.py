from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table
from pytest import MonkeyPatch

from app.api.workbooks import get_insight_generator
from app.main import app
from app.services.analysis_strategy import AnalysisDepth
from app.services.insight_generator import WorkbookInsight, WorkbookInsightReport

client = TestClient(app)


class StubInsightGenerator:
    def __init__(self) -> None:
        self.requested_depth = AnalysisDepth.AUTO

    async def generate(
        self,
        summary: object,
        depth: AnalysisDepth = AnalysisDepth.AUTO,
    ) -> WorkbookInsightReport:
        self.requested_depth = depth
        return WorkbookInsightReport(
            overview="2개 시트로 구성된 워크북입니다.",
            insights=[
                WorkbookInsight(
                    title="수식 검토 필요",
                    description="매출현황 시트에 합계 수식이 있습니다.",
                    category="formula",
                    severity="info",
                    evidence=["매출현황!D2 = SUM(B2:C2)"],
                    recommendation="합계 범위를 확인하세요.",
                )
            ],
            limitations=["실제 셀 값의 의미는 분석하지 않았습니다."],
        )


def create_workbook_file() -> bytes:
    workbook = Workbook()
    sales_sheet = workbook.active
    sales_sheet.title = "매출현황"
    sales_sheet.append(["상품", "1월", "2월", "합계"])
    sales_sheet.append(["노트북", 10, 20, "=SUM(B2:C2)"])
    sales_sheet.append(["모니터", 5, 15, "=SUM(B3:C3)"])
    sales_sheet.add_table(Table(displayName="SalesTable", ref="A1:D3"))

    chart = BarChart()
    chart.add_data(Reference(sales_sheet, min_col=2, max_col=3, min_row=1, max_row=3))
    sales_sheet.add_chart(chart, "F2")

    summary_sheet = workbook.create_sheet("요약")
    summary_sheet["A1"] = "완료"
    summary_sheet["A2"] = "=매출현황!D2"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def create_workbook_with_system_sheets() -> bytes:
    workbook = Workbook()
    business_sheet = workbook.active
    business_sheet.title = "업무 데이터"
    business_sheet["A1"] = "분석 대상"

    hidden_sheet = workbook.create_sheet("숨김 계산")
    hidden_sheet["A1"] = "화면에 표시하지 않음"
    hidden_sheet.sheet_state = "hidden"

    add_in_sheet = workbook.create_sheet("__snlofficequeries")
    add_in_sheet["A1"] = "애드인 캐시"

    cache_sheet = workbook.create_sheet("CIOHiddenCacheSheet")
    cache_sheet["A1"] = "시스템 캐시"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_returns_workbook_summary() -> None:
    response = client.post(
        "/api/v1/workbooks/summary",
        files={
            "file": (
                "sales.xlsx",
                create_workbook_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    result = response.json()
    assert result["filename"] == "sales.xlsx"
    assert result["sheet_count"] == 2
    assert result["total_sheet_count"] == 2
    assert result["excluded_sheet_count"] == 0
    assert result["excluded_sheets"] == []

    sales = result["sheets"][0]
    assert sales["name"] == "매출현황"
    assert sales["analysis_inclusion"]["decision"] == "include"
    assert sales["analysis_inclusion"]["reason_code"] == "business_worksheet"
    assert sales["formula_count"] == 2
    assert sales["formulas"][0] == {
        "cell": "D2",
        "formula": "=SUM(B2:C2)",
        "references": ["B2:C2"],
        "cached_value": None,
        "role": "calculation",
    }
    formula_cell = sales["regions"][0]["preview_rows"][1][3]
    assert formula_cell["address"] == "D2"
    assert formula_cell["value"] is None
    assert formula_cell["formula"] == "=SUM(B2:C2)"
    assert formula_cell["cached_value"] is None
    assert formula_cell["number_format"] == "General"
    assert formula_cell["bold"] is False
    assert formula_cell["merged"] is False
    assert formula_cell["semantic"] is None
    assert sales["regions"][0]["title"] == "상품"
    assert sales["regions"][0]["semantic"] is None
    assert sales["regions"][0]["analysis_inclusion"]["decision"] == "include"
    assert (
        sales["regions"][0]["analysis_inclusion"]["reason_code"]
        == "populated_business_region"
    )
    assert sales["regions"][0]["row_count"] == 3
    assert sales["regions"][0]["column_count"] == 4
    assert sales["regions"][0]["header_paths"][0] == {
        "column": "A",
        "labels": ["상품"],
    }
    assert sales["tables"][0]["name"] == "SalesTable"
    assert sales["tables"][0]["reference"] == "A1:D3"
    assert sales["tables"][0]["headers"] == ["상품", "1월", "2월", "합계"]
    assert sales["charts"][0]["chart_type"] == "BarChart"
    assert sales["charts"][0]["anchor_cell"] == "F2"
    assert sales["charts"][0]["series"][0]["value_samples"] == ["1월", 10, 5]

    summary = result["sheets"][1]
    assert summary["name"] == "요약"
    assert summary["regions"][0]["preview_rows"][0][0]["value"] == "완료"

    dependencies = result["dependency_summary"]
    assert dependencies["formula_node_count"] == 3
    assert dependencies["edge_count"] == 3
    assert dependencies["cross_sheet_edge_count"] == 1
    assert dependencies["cluster_count"] == 2
    assert dependencies["clusters"][0]["sheet_names"] == ["매출현황", "요약"]
    assert dependencies["clusters"][0]["edges"][-1] == {
        "source": "매출현황!D2",
        "target": "요약!A2",
        "reference": "매출현황!D2",
        "cross_sheet": True,
    }


def test_excludes_hidden_and_add_in_cache_sheets() -> None:
    response = client.post(
        "/api/v1/workbooks/summary",
        files={
            "file": (
                "system-sheets.xlsx",
                create_workbook_with_system_sheets(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    result = response.json()
    assert result["sheet_count"] == 1
    assert result["total_sheet_count"] == 4
    assert result["excluded_sheet_count"] == 3
    assert [sheet["name"] for sheet in result["sheets"]] == ["업무 데이터"]
    assert [sheet["name"] for sheet in result["excluded_sheets"]] == [
        "숨김 계산",
        "__snlofficequeries",
        "CIOHiddenCacheSheet",
    ]
    assert [
        sheet["analysis_inclusion"]["reason_code"]
        for sheet in result["excluded_sheets"]
    ] == [
        "hidden_worksheet",
        "addin_cache_worksheet",
        "system_cache_worksheet",
    ]


def test_rejects_unsupported_extension() -> None:
    response = client.post(
        "/api/v1/workbooks/summary",
        files={"file": ("sales.csv", b"name,value", "text/csv")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == ".xlsx 또는 .xlsm 파일만 업로드할 수 있습니다."


def test_rejects_invalid_excel_file() -> None:
    response = client.post(
        "/api/v1/workbooks/summary",
        files={"file": ("sales.xlsx", b"not-an-excel-file")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "올바른 Excel 파일이 아닙니다."


def test_rejects_empty_file() -> None:
    response = client.post(
        "/api/v1/workbooks/summary",
        files={"file": ("sales.xlsx", b"")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "빈 파일은 업로드할 수 없습니다."


def test_rejects_file_exceeding_size_limit(monkeypatch: MonkeyPatch) -> None:
    monkeypatch.setattr("app.api.workbooks.MAX_FILE_SIZE_BYTES", 5)

    response = client.post(
        "/api/v1/workbooks/summary",
        files={"file": ("sales.xlsx", b"123456")},
    )

    assert response.status_code == 413
    assert response.json()["detail"] == "파일 크기는 50MB를 초과할 수 없습니다."


def test_returns_structured_workbook_insights() -> None:
    insight_generator = StubInsightGenerator()
    app.dependency_overrides[get_insight_generator] = lambda: insight_generator

    try:
        response = client.post(
            "/api/v1/workbooks/insights",
            data={"depth": "PRECISE"},
            files={
                "file": (
                    "sales.xlsx",
                    create_workbook_file(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200
    assert insight_generator.requested_depth == AnalysisDepth.PRECISE
    result = response.json()
    assert result["workbook"]["filename"] == "sales.xlsx"
    assert result["report"] == {
        "overview": "2개 시트로 구성된 워크북입니다.",
        "insights": [
            {
                "title": "수식 검토 필요",
                "description": "매출현황 시트에 합계 수식이 있습니다.",
                "category": "formula",
                "severity": "info",
                "evidence": ["매출현황!D2 = SUM(B2:C2)"],
                "recommendation": "합계 범위를 확인하세요.",
            }
        ],
        "limitations": ["실제 셀 값의 의미는 분석하지 않았습니다."],
    }


def test_rejects_unknown_analysis_depth() -> None:
    app.dependency_overrides[get_insight_generator] = StubInsightGenerator

    try:
        response = client.post(
            "/api/v1/workbooks/insights",
            data={"depth": "UNKNOWN"},
            files={
                "file": (
                    "sales.xlsx",
                    create_workbook_file(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 422


def test_returns_service_unavailable_without_openai_api_key(
    monkeypatch: MonkeyPatch,
) -> None:
    monkeypatch.delenv("OPENAI_API_KEY", raising=False)
    monkeypatch.setattr("app.services.insight_generator.load_dotenv", lambda _: False)

    response = client.post(
        "/api/v1/workbooks/insights",
        files={
            "file": (
                "sales.xlsx",
                create_workbook_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 503
    assert response.json()["detail"] == (
        "OPENAI_API_KEY가 설정되지 않았습니다. AI/.env 파일을 확인하세요."
    )
