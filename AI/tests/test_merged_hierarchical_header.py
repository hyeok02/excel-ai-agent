from pathlib import Path

from openpyxl import Workbook

from app.services.region_detector import detect_regions
from app.services.workbook_details import summarize_regions
from app.services.workbook_parser import parse_workbook

FIXTURE = Path(__file__).parent / "fixtures/semantic/semantic_hierarchical_headers.xlsx"


def test_detects_merged_hierarchical_header_from_fixture() -> None:
    summary = parse_workbook(FIXTURE.name, FIXTURE.read_bytes())
    header = next(
        region
        for region in summary.sheets[0].regions
        if region.semantic and region.semantic.role.value == "header"
    )

    assert (header.start_cell, header.end_cell) == ("A4", "E5")
    assert any(
        reason.code == "merged_hierarchical_header"
        for reason in header.semantic.reasons
    )
    assert [(path.column, path.labels) for path in header.header_paths] == [
        ("A", ["부서"]),
        ("B", ["매출", "계획"]),
        ("C", ["매출", "실적"]),
        ("D", ["비용", "계획"]),
        ("E", ["비용", "실적"]),
    ]


def test_resolves_horizontal_and_vertical_merged_header_anchors() -> None:
    workbook = _hierarchical_workbook()
    worksheet = workbook.active
    regions = detect_regions(worksheet)
    summaries = summarize_regions(worksheet, regions)
    header = next(item for item in summaries if item.semantic.role.value == "header")

    assert [(path.column, path.labels) for path in header.header_paths] == [
        ("A", ["부서"]),
        ("B", ["매출", "계획"]),
        ("C", ["매출", "실적"]),
    ]
    workbook.close()


def test_does_not_treat_full_width_title_merge_as_hierarchical_header() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells("A1:C1")
    worksheet["A1"] = "분기별 실적"
    worksheet.append(["부서", "계획", "실적"])
    worksheet.append(["영업", 100, 110])
    worksheet.append(["개발", 90, 105])

    regions = detect_regions(worksheet)
    reasons = [reason.code for region in regions for reason in region.semantic.reasons]

    assert "merged_hierarchical_header" not in reasons
    assert [region.semantic.role.value for region in regions][-2:] == ["header", "data"]
    workbook.close()


def _hierarchical_workbook() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells("A1:A2")
    worksheet.merge_cells("B1:C1")
    worksheet["A1"] = "부서"
    worksheet["B1"] = "매출"
    worksheet["B2"] = "계획"
    worksheet["C2"] = "실적"
    worksheet.append(["영업", 100, 110])
    worksheet.append(["개발", 90, 105])
    return workbook
