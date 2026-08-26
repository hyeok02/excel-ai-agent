from openpyxl import Workbook

from app.services.region_detector import detect_regions
from app.services.workbook_details import summarize_regions
from app.services.workbook_details.header_path_builder import normalize_path


def test_generates_full_five_level_header_paths() -> None:
    workbook = _five_level_workbook()
    header = _header_summary(workbook)

    assert (header.start_cell, header.end_cell) == ("A1", "E5")
    assert [(path.column, path.labels) for path in header.header_paths] == [
        ("A", ["구분"]),
        ("B", ["인건비", "국내", "정규직", "기본급", "계획"]),
        ("C", ["인건비", "국내", "계약직", "성과급", "실적"]),
        ("D", ["인건비", "해외", "정규직", "기본급", "계획"]),
        ("E", ["인건비", "해외", "계약직", "성과급", "실적"]),
    ]
    workbook.close()


def test_generates_paths_beyond_the_twelfth_column() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells("A1:A3")
    worksheet.merge_cells("B1:M1")
    worksheet["A1"] = "구분"
    worksheet["B1"] = "인건비"
    for column in range(2, 14):
        worksheet.cell(2, column, f"지역 {column - 1}")
        worksheet.cell(3, column, "실적")
    worksheet.append(["1월", *range(1, 13)])
    worksheet.append(["2월", *range(2, 14)])

    header = _header_summary(workbook)

    assert len(header.header_paths) == 13
    assert header.header_paths[-1].column == "M"
    assert header.header_paths[-1].labels == ["인건비", "지역 12", "실적"]
    workbook.close()


def test_normalizes_blank_repeated_and_spaced_levels() -> None:
    assert normalize_path([" 인건비 ", None, "인건비", "국내   정규직"]) == [
        "인건비",
        "국내 정규직",
    ]


def _header_summary(workbook: Workbook):
    worksheet = workbook.active
    regions = detect_regions(worksheet)
    summaries = summarize_regions(worksheet, regions)
    return next(item for item in summaries if item.semantic.role.value == "header")


def _five_level_workbook() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells("A1:A5")
    worksheet.merge_cells("B1:E1")
    worksheet.merge_cells("B2:C2")
    worksheet.merge_cells("D2:E2")
    worksheet["A1"], worksheet["B1"] = "구분", "인건비"
    worksheet["B2"], worksheet["D2"] = "국내", "해외"
    rows = [
        ["정규직", "계약직", "정규직", "계약직"],
        ["기본급", "성과급", "기본급", "성과급"],
        ["계획", "실적", "계획", "실적"],
    ]
    for row_number, labels in enumerate(rows, start=3):
        for column, label in enumerate(labels, start=2):
            worksheet.cell(row_number, column, label)
    worksheet.append(["영업", 100, 110, 90, 95])
    worksheet.append(["개발", 120, 125, 105, 115])
    return workbook
