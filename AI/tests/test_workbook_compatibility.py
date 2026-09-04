from copy import copy
from io import BytesIO
from zipfile import ZipFile

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.agent.query import build_workbook_data_index
from app.agent.writeback import apply_writeback
from app.agent.writeback.models import WritebackChange
from app.main import app
from app.services import workbook_loading
from app.services.workbook_loading import (
    STYLE_ERROR, InvalidWorkbookError,
    load_workbook_for_reading, prepare_workbook_for_reading,
)
from app.services.workbook_parser import parse_workbook
from tests.support.compatibility_workbook import (
    compatibility_workbook, replace_part, standard_workbook,
)
from tests.support.workbook_api_fixtures import upload


def test_standard_workbook_uses_original_bytes_without_repacking() -> None:
    original = standard_workbook()
    assert prepare_workbook_for_reading(original) is original


def test_fallback_styles_preserve_all_indices_values_formulas_and_dates() -> None:
    original = standard_workbook()
    compatible = compatibility_workbook(original)
    with pytest.raises(IndexError):
        load_workbook(BytesIO(compatible))

    assert parse_workbook("sales.xlsx", compatible) == parse_workbook("sales.xlsx", original)
    assert build_workbook_data_index("sales.xlsx", compatible) == build_workbook_data_index(
        "sales.xlsx", original
    )
    expected = load_workbook(BytesIO(original))
    actual = load_workbook_for_reading(compatible)
    try:
        for address in ("B2", "B3", "B4", "D2"):
            before, after = expected["매출현황"][address], actual["매출현황"][address]
            assert before.value == after.value
            assert before.style_id == after.style_id
            assert before.font == copy(after.font)
            assert before.alignment == copy(after.alignment)
            assert before.number_format == after.number_format
        assert str(expected["매출현황"].merged_cells) == str(actual["매출현황"].merged_cells)
    finally:
        expected.close()
        actual.close()


def test_only_temporary_styles_part_changes() -> None:
    original = compatibility_workbook()
    prepared = prepare_workbook_for_reading(original)
    with ZipFile(BytesIO(original)) as source, ZipFile(BytesIO(prepared)) as view:
        assert source.namelist() == view.namelist()
        assert b"AlternateContent" in source.read("xl/styles.xml")
        assert b"AlternateContent" not in view.read("xl/styles.xml")
        for name in source.namelist():
            if name != "xl/styles.xml":
                assert source.read(name) == view.read(name)


@pytest.mark.parametrize("invalid", ["missing", "multiple", "foreign", "duplicate"])
def test_unsafe_fallback_fails_explicitly_without_dropping_style_slots(invalid) -> None:
    with pytest.raises(InvalidWorkbookError, match="Excel 호환 서식"):
        parse_workbook("sales.xlsx", compatibility_workbook(invalid=invalid))


def test_index_out_of_bounds_becomes_user_readable_validation_error() -> None:
    original = standard_workbook()
    with ZipFile(BytesIO(original)) as archive:
        xml = archive.read("xl/worksheets/sheet1.xml")
    broken = replace_part(original, "xl/worksheets/sheet1.xml", xml.replace(b's="1"', b's="99999"'))
    response = TestClient(app).post("/api/v1/workbooks/summary", files=upload("broken.xlsx", broken))
    assert response.status_code == 400
    assert response.json()["detail"] == STYLE_ERROR


def test_entities_in_compatibility_styles_are_rejected() -> None:
    xml = b'<!DOCTYPE styleSheet [<!ENTITY x "AlternateContent">]><styleSheet>&x;</styleSheet>'
    content = replace_part(standard_workbook(), "xl/styles.xml", xml)
    with pytest.raises(InvalidWorkbookError, match="Excel 호환 서식"):
        prepare_workbook_for_reading(content)


def test_pair_closes_first_workbook_if_second_load_fails(monkeypatch) -> None:
    class Book:
        closed = False

        def close(self):
            self.closed = True

    first = Book()
    def load(content, **kwargs):
        if kwargs["data_only"]:
            raise InvalidWorkbookError(STYLE_ERROR)
        return first

    monkeypatch.setattr(workbook_loading, "_load_prepared", load)
    with pytest.raises(InvalidWorkbookError):
        workbook_loading.load_workbook_pair(standard_workbook())
    assert first.closed


@pytest.mark.parametrize("filename", ["sales.xlsx", "sales.xlsm"])
def test_writeback_preserves_original_compatibility_styles_and_other_zip_parts(filename) -> None:
    original = compatibility_workbook()
    change = WritebackChange(
        sheet_name="매출현황", reference="B2", old_value=10, new_value=12,
        reason="사용자가 승인한 정정", value_type="number",
    )
    modified, manifest = apply_writeback(filename, original, [change])
    assert manifest.verified
    before = load_workbook_for_reading(original)
    after = load_workbook_for_reading(modified)
    try:
        assert before["매출현황"]["B2"].value == 10
        assert after["매출현황"]["B2"].value == 12
        assert before["매출현황"]["B2"].style_id == after["매출현황"]["B2"].style_id
    finally:
        before.close()
        after.close()
    with ZipFile(BytesIO(original)) as source, ZipFile(BytesIO(modified)) as result:
        assert source.read("xl/styles.xml") == result.read("xl/styles.xml")
        assert b"AlternateContent" in result.read("xl/styles.xml")
        for part in source.namelist():
            if part not in {"xl/worksheets/sheet1.xml", "xl/workbook.xml"}:
                assert source.read(part) == result.read(part)
