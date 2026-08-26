from datetime import date

from openpyxl import Workbook

from app.services.region_detector import detect_regions
from app.services.workbook_details import build_column_schemas, summarize_regions


def test_recognizes_units_and_standard_fields() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["날짜", "매출액", "영업이익률", "인원수", "판매수량"])
    sheet.append([date(2026, 1, 1), 1200000, 0.21, 18, 40])
    sheet.append([date(2026, 2, 1), 1350000, 0.24, 20, 43])
    for row in range(2, 4):
        sheet.cell(row, 1).number_format = "yyyy-mm-dd"
        sheet.cell(row, 2).number_format = '₩#,##0'
        sheet.cell(row, 3).number_format = "0.0%"
        sheet.cell(row, 4).number_format = '0 "명"'
        sheet.cell(row, 5).number_format = '0 "개"'

    schemas = _schemas(sheet)

    assert [item.standard_field for item in schemas] == [
        "period",
        "revenue",
        "profit_margin",
        "headcount",
        "quantity",
    ]
    assert [item.unit_type for item in schemas] == [
        "date",
        "currency",
        "percentage",
        "headcount",
        "quantity",
    ]
    assert schemas[1].unit_label == "KRW"
    assert schemas[2].unit_label == "%"
    workbook.close()


def test_normalizes_revenue_synonyms_to_same_standard_field() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Revenue", "Sales", "매출액"])
    sheet.append([100, 110, 120])
    sheet.append([130, 140, 150])

    schemas = _schemas(sheet)

    assert [item.standard_field for item in schemas] == [
        "revenue",
        "revenue",
        "revenue",
    ]
    assert all(item.confidence >= 0.7 for item in schemas)
    workbook.close()


def _schemas(sheet):
    regions = detect_regions(sheet)
    summaries = summarize_regions(sheet, regions, sheet)
    return build_column_schemas(sheet, sheet, regions, summaries)
