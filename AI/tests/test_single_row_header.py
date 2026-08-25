from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.region_detector import detect_regions
from app.services.workbook_details import summarize_regions


def test_detects_unstyled_single_row_header_from_data_transition() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["지역", "1월", "2월", "합계"])
    worksheet.append(["서울", 120, 135, 255])
    worksheet.append(["부산", 90, 95, 185])
    worksheet.append(["대구", 75, 80, 155])

    regions = detect_regions(worksheet)

    assert _roles(regions) == ["header", "data"]
    assert (regions[0].start_cell, regions[0].end_cell) == ("A1", "D1")
    assert regions[0].semantic.confidence >= 0.65
    assert any(
        reason.code == "single_row_header" for reason in regions[0].semantic.reasons
    )
    workbook.close()


def test_separates_title_immediately_above_single_header() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells("A1:C1")
    worksheet["A1"] = "분기별 매출 현황"
    worksheet["A1"].font = Font(bold=True)
    worksheet.append(["부서", "매출", "비용"])
    worksheet.append(["영업", 300, 180])
    worksheet.append(["개발", 250, 210])

    regions = detect_regions(worksheet)

    assert _roles(regions) == ["title", "header", "data"]
    assert [(region.start_cell, region.end_cell) for region in regions] == [
        ("A1", "C1"),
        ("A2", "C2"),
        ("A3", "C4"),
    ]
    workbook.close()


def test_recognizes_styled_header_with_numeric_period_labels() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["구분", 2025, 2026, "증감률"])
    worksheet.append(["매출", 100, 120, 0.2])
    worksheet.append(["비용", 70, 76, 0.09])
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCE8FF")

    regions = detect_regions(worksheet)
    summaries = summarize_regions(worksheet, regions)

    assert _roles(regions) == ["header", "data"]
    assert [path.labels for path in summaries[0].header_paths] == [
        ["구분"],
        ["2025"],
        ["2026"],
        ["증감률"],
    ]
    workbook.close()


def test_does_not_invent_header_for_headerless_numeric_rows() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["서울", 120, 135])
    worksheet.append(["부산", 90, 95])
    worksheet.append(["대구", 75, 80])

    regions = detect_regions(worksheet)

    assert "header" not in _roles(regions)
    workbook.close()


def _roles(regions: list[object]) -> list[str]:
    return [region.semantic.role.value for region in regions]
