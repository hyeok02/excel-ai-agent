from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table

from app.services.analysis_strategy import AnalysisDepth
from app.services.insight_generator import WorkbookInsight, WorkbookInsightReport

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class StubInsightGenerator:
    def __init__(self) -> None:
        self.requested_depth = AnalysisDepth.AUTO

    async def generate(
        self,
        summary: object,
        depth: AnalysisDepth = AnalysisDepth.AUTO,
    ) -> WorkbookInsightReport:
        self.requested_depth = depth
        return WorkbookInsightReport(
            overview="2개 시트로 구성된 워크북입니다.",
            insights=[
                WorkbookInsight(
                    title="수식 검토 필요",
                    description="매출현황 시트에 합계 수식이 있습니다.",
                    category="formula",
                    severity="info",
                    evidence=["매출현황!D2 = SUM(B2:C2)"],
                    recommendation="합계 범위를 확인하세요.",
                )
            ],
            limitations=["실제 셀 값의 의미는 분석하지 않았습니다."],
        )


def create_workbook_file() -> bytes:
    workbook = Workbook()
    sales = workbook.active
    sales.title = "매출현황"
    sales.append(["상품", "1월", "2월", "합계"])
    sales.append(["노트북", 10, 20, "=SUM(B2:C2)"])
    sales.append(["모니터", 5, 15, "=SUM(B3:C3)"])
    sales.add_table(Table(displayName="SalesTable", ref="A1:D3"))
    chart = BarChart()
    chart.add_data(Reference(sales, min_col=2, max_col=3, min_row=1, max_row=3))
    sales.add_chart(chart, "F2")
    summary = workbook.create_sheet("요약")
    summary["A1"] = "완료"
    summary["A2"] = "=매출현황!D2"
    return _save(workbook)


def create_workbook_with_system_sheets() -> bytes:
    workbook = Workbook()
    business = workbook.active
    business.title = "업무 데이터"
    business["A1"] = "분석 대상"
    hidden = workbook.create_sheet("숨김 계산")
    hidden["A1"] = "화면에 표시하지 않음"
    hidden.sheet_state = "hidden"
    workbook.create_sheet("__snlofficequeries")["A1"] = "애드인 캐시"
    workbook.create_sheet("CIOHiddenCacheSheet")["A1"] = "시스템 캐시"
    return _save(workbook)


def upload(filename: str, content: bytes, mime: str = EXCEL_MIME) -> dict:
    return {"file": (filename, content, mime)}


def _save(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
