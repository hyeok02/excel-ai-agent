"""Synthetic compatibility styles: no user workbook contents in the repository."""

from copy import copy, deepcopy
from datetime import datetime
from io import BytesIO
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
HS_NS = "http://schemas.haansoft.com/office/spreadsheet/8.0"


def standard_workbook() -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = "매출현황"
    sheet.append(["상품", "1월", "2월", "합계"])
    sheet.append(["노트북", 10, 20, "=SUM(B2:C2)"])
    sheet.append(["모니터", 5, 15, "=SUM(B3:C3)"])
    sheet.append(["기준일", datetime(2026, 9, 4)])
    sheet["B2"].font = Font(name="맑은 고딕", bold=True, color="FF112233")
    sheet["B2"].alignment = Alignment(horizontal="right", wrap_text=True)
    sheet["B2"].number_format = "0.00"
    sheet["B3"].font = Font(italic=True, color="FF445566")
    sheet["B4"].number_format = "yyyy-mm-dd"
    sheet.merge_cells("A5:D5")
    sheet["A5"] = "병합 표 제목"
    book.create_sheet("요약")["A1"] = "=매출현황!D2"
    stream = BytesIO()
    book.save(stream)
    book.close()
    return stream.getvalue()


def replace_part(content: bytes, part: str, replacement: bytes) -> bytes:
    stream = BytesIO()
    with ZipFile(BytesIO(content)) as source, ZipFile(stream, "w") as target:
        for info in source.infolist():
            target.writestr(
                copy(info), replacement if info.filename == part else source.read(info.filename)
            )
    return stream.getvalue()


def compatibility_workbook(content: bytes | None = None, *, invalid: str | None = None) -> bytes:
    content = content if content is not None else standard_workbook()
    with ZipFile(BytesIO(content)) as archive:
        styles = ET.fromstring(archive.read("xl/styles.xml"))
    for name in ("fonts", "cellXfs"):
        parent = styles.find(f"{{{MAIN_NS}}}{name}")
        # Wrap an interior slot as well as the last one, retaining original IDs.
        for index, item in enumerate(list(parent)):
            if index == 0:
                continue
            wrapper = ET.Element(f"{{{MC_NS}}}AlternateContent")
            choice = ET.SubElement(wrapper, f"{{{MC_NS}}}Choice", {"Requires": "hs"})
            extended = deepcopy(item)
            extended.set(f"{{{HS_NS}}}applyExtension", "1")
            choice.append(extended)
            if invalid != "missing":
                fallback = ET.SubElement(wrapper, f"{{{MC_NS}}}Fallback")
                fallback.append(deepcopy(item))
                if invalid == "multiple":
                    fallback.append(deepcopy(item))
                if invalid == "foreign":
                    fallback[0].tag = "{urn:unrelated}xf"
                if invalid == "duplicate":
                    wrapper.append(deepcopy(fallback))
            parent.remove(item)
            parent.insert(index, wrapper)
    return replace_part(content, "xl/styles.xml", ET.tostring(styles))
