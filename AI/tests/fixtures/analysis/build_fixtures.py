"""도메인이 서로 다른 회귀 픽스처 워크북을 생성한다.

바이너리 .xlsx만 커밋하면 리뷰에서 내용을 확인할 수 없으므로, 픽스처는
이 스크립트로 재생성 가능하게 유지한다.

    cd AI && python -m tests.fixtures.analysis.build_fixtures
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

DIRECTORY = Path(__file__).parent
MONTHS = ("2025-01-01", "2025-02-01", "2025-03-01", "2025-04-01", "2025-05-01", "2025-06-01")


def _title(sheet, text: str) -> None:
    sheet["A1"] = text
    sheet["A1"].font = Font(bold=True)


def build_cost_close_no_identity() -> Workbook:
    """식별 행이 없는 원가 마감표. 분석 대상 이름을 뽑을 수 없어야 한다."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "원가마감"
    _title(sheet, "2025년 상반기 원가 마감")
    sheet["A2"] = "단위: 천원"
    sheet.append([])
    sheet.append(["기준월", "재료비", "가공비", "총원가"])
    material = (4200, 4310, 4480, 4610, 4720, 4980)
    processing = (1800, 1770, 1810, 1690, 1640, 1580)
    for month, first, second in zip(MONTHS, material, processing):
        sheet.append([month, first, second, first + second])
    return workbook


def build_equipment_snapshot() -> Workbook:
    """기간 추이가 없는 설비 점검 스냅샷. 변화 지표가 나오지 않아야 한다."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "설비점검"
    _title(sheet, "설비 점검 현황")
    sheet.append([])
    sheet.append(["설비명", "2공장 압출 1호기"])
    sheet.append([])
    sheet.append(["설비코드", "점검항목", "측정값", "기준값"])
    rows = (
        ("EX-101", "진동", 3.2, 5.0),
        ("EX-102", "베어링 온도", 68.4, 80.0),
        ("EX-103", "토출 압력", 12.7, 15.0),
        ("EX-104", "스크류 회전수", 118.0, 130.0),
    )
    for row in rows:
        sheet.append(list(row))
    return workbook


def build_yield_report_english() -> Workbook:
    """한국어가 전혀 없는 워크북. 언어 중립성을 확인한다."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Yield"
    _title(sheet, "Monthly Yield Report")
    sheet.append([])
    sheet.append(["Line", "Extrusion Line B"])
    sheet.append([])
    sheet.append(["Period", "Output", "Defects", "Yield Rate"])
    output = (12000, 12400, 11800, 12900, 13200, 13500)
    defects = (240, 260, 300, 258, 231, 189)
    for month, made, bad in zip(MONTHS, output, defects):
        sheet.append([month, made, bad, round((made - bad) / made * 100, 2)])
    return workbook


def build_energy_mixed_layout() -> Workbook:
    """소계가 중간에 끼고 헤더가 반복되는 실무형 배치."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "에너지"
    _title(sheet, "공장별 에너지 사용량")
    sheet.merge_cells("A1:D1")
    sheet["A2"] = "단위: kWh"
    sheet.append([])
    sheet.append(["기간", "1공장", "2공장", "합계"])
    first_half = ((MONTHS[0], 82000, 61000), (MONTHS[1], 84500, 60200), (MONTHS[2], 86100, 62400))
    for month, plant_one, plant_two in first_half:
        sheet.append([month, plant_one, plant_two, plant_one + plant_two])
    sheet.append(["소계", 252600, 183600, 436200])
    sheet.append([])
    sheet.append(["기간", "1공장", "2공장", "합계"])
    second_half = ((MONTHS[3], 88300, 63100), (MONTHS[4], 90200, 64800), (MONTHS[5], 91700, 65500))
    for month, plant_one, plant_two in second_half:
        sheet.append([month, plant_one, plant_two, plant_one + plant_two])
    guide = workbook.create_sheet("사용안내")
    guide["A1"] = "본 자료는 월별 검침값을 옮겨 적은 것입니다."
    guide["A2"] = "검침 누락 월은 직전 월 값을 사용합니다."
    return workbook


BUILDERS = {
    "cost_close_no_identity.xlsx": build_cost_close_no_identity,
    "equipment_snapshot_no_trend.xlsx": build_equipment_snapshot,
    "yield_report_english_only.xlsx": build_yield_report_english,
    "energy_mixed_layout.xlsx": build_energy_mixed_layout,
}


def main() -> None:
    for name, builder in BUILDERS.items():
        workbook = builder()
        workbook.save(DIRECTORY / name)
        workbook.close()
        print(f"생성: {name}")


if __name__ == "__main__":
    main()
