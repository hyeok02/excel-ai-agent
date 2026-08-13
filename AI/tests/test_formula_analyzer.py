from openpyxl import Workbook

from app.services.formula_analyzer import FormulaAnalysis, analyze_formulas


def test_extracts_cell_and_range_references() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["D2"] = "=SUM(B2:C2)+$A$1"

    assert analyze_formulas(worksheet) == [
        FormulaAnalysis(
            cell="D2",
            formula="=SUM(B2:C2)+$A$1",
            references=["B2:C2", "$A$1"],
        )
    ]
    workbook.close()


def test_preserves_cross_sheet_reference() -> None:
    workbook = Workbook()
    source = workbook.active
    source.title = "기준 데이터"
    summary = workbook.create_sheet("요약")
    summary["A1"] = "='기준 데이터'!$B$2"

    assert analyze_formulas(summary) == [
        FormulaAnalysis(
            cell="A1",
            formula="='기준 데이터'!$B$2",
            references=["'기준 데이터'!$B$2"],
        )
    ]
    workbook.close()


def test_ignores_named_ranges_and_removes_duplicate_references() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["B1"] = "=IF(A1>0,A1+NamedRange,0)"

    assert analyze_formulas(worksheet) == [
        FormulaAnalysis(
            cell="B1",
            formula="=IF(A1>0,A1+NamedRange,0)",
            references=["A1"],
        )
    ]
    workbook.close()
