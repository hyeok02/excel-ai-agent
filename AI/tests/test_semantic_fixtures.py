import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


FIXTURE_DIRECTORY = Path(__file__).parent / "fixtures" / "semantic"
ALLOWED_SHEET_DECISIONS = {"analyze", "metadata_only", "exclude"}
ALLOWED_SHEET_ROLES = {
    "business_data",
    "business_model",
    "instruction",
    "system_cache",
}
ALLOWED_REGION_DECISIONS = {"analyze", "context", "exclude"}
ALLOWED_REGION_ROLES = {
    "title",
    "unit",
    "header",
    "data",
    "total",
    "source_note",
    "rule_note",
    "input",
    "calculation",
    "output",
    "instruction",
    "warning",
    "system_cache",
}


def test_semantic_fixture_manifest_is_complete() -> None:
    manifest = _load_json(FIXTURE_DIRECTORY / "manifest.json")

    assert manifest["schema_version"] == 1
    assert len(manifest["fixtures"]) == 3

    workbook_names = [fixture["workbook"] for fixture in manifest["fixtures"]]
    expectation_names = [fixture["expectation"] for fixture in manifest["fixtures"]]
    assert len(workbook_names) == len(set(workbook_names))
    assert len(expectation_names) == len(set(expectation_names))

    for fixture in manifest["fixtures"]:
        assert fixture["coverage"]
        assert (FIXTURE_DIRECTORY / fixture["workbook"]).is_file()
        assert (FIXTURE_DIRECTORY / fixture["expectation"]).is_file()


def test_semantic_expectations_reference_valid_workbook_ranges() -> None:
    manifest = _load_json(FIXTURE_DIRECTORY / "manifest.json")

    for fixture in manifest["fixtures"]:
        workbook_path = FIXTURE_DIRECTORY / fixture["workbook"]
        expectation_path = FIXTURE_DIRECTORY / fixture["expectation"]
        expectation = _load_json(expectation_path)
        workbook = load_workbook(workbook_path, data_only=False)

        try:
            assert expectation["schema_version"] == 1
            assert expectation["workbook"] == fixture["workbook"]
            assert expectation["workbook_intent"].strip()
            assert {sheet["name"] for sheet in expectation["sheets"]} == set(
                workbook.sheetnames
            )
            assert any(
                sheet["decision"] == "analyze"
                for sheet in expectation["sheets"]
            )

            for expected_sheet in expectation["sheets"]:
                worksheet = workbook[expected_sheet["name"]]
                assert expected_sheet["decision"] in ALLOWED_SHEET_DECISIONS
                assert expected_sheet["sheet_role"] in ALLOWED_SHEET_ROLES
                assert expected_sheet["reason"].strip()
                assert expected_sheet["regions"]
                assert _formula_count(worksheet) == expected_sheet[
                    "expected_formula_count"
                ]

                region_decisions = {
                    region["decision"] for region in expected_sheet["regions"]
                }
                if expected_sheet["decision"] == "analyze":
                    assert "analyze" in region_decisions
                elif expected_sheet["decision"] == "metadata_only":
                    assert region_decisions <= {"context"}
                else:
                    assert region_decisions == {"exclude"}

                for region in expected_sheet["regions"]:
                    assert region["role"] in ALLOWED_REGION_ROLES
                    assert region["decision"] in ALLOWED_REGION_DECISIONS
                    assert region["reason"].strip()
                    _assert_valid_range(worksheet, region["range"])
                    for unit in region.get("units", []):
                        assert unit["unit"].strip()
                        assert unit["source"] in {"label", "number_format"}
                        _assert_valid_range(worksheet, unit["range"])

            for reference in expectation["reference_answers"]:
                assert reference["question"].strip()
                assert reference["answer_cells"]
                for address in reference["answer_cells"]:
                    _assert_valid_qualified_range(workbook, address)
                for address in reference["supporting_cells"]:
                    _assert_valid_qualified_range(workbook, address)
        finally:
            workbook.close()


def _load_json(path: Path) -> dict[str, object]:
    with path.open(encoding="utf-8") as file:
        return json.load(file)


def _formula_count(worksheet: object) -> int:
    return sum(
        1
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )


def _assert_valid_qualified_range(workbook: object, address: str) -> None:
    sheet_name, cell_range = address.rsplit("!", maxsplit=1)
    assert sheet_name in workbook.sheetnames
    _assert_valid_range(workbook[sheet_name], cell_range)


def _assert_valid_range(worksheet: object, cell_range: str) -> None:
    min_column, min_row, max_column, max_row = range_boundaries(cell_range)
    assert min_row >= 1
    assert min_column >= 1
    assert max_row <= worksheet.max_row
    assert max_column <= worksheet.max_column
    assert any(
        worksheet.cell(row=row, column=column).value is not None
        for row in range(min_row, max_row + 1)
        for column in range(min_column, max_column + 1)
    )
