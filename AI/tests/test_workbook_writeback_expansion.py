import json
from datetime import date, datetime
from io import BytesIO
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.api.workbook_writebacks import get_writeback_generator
from app.main import app
from tests.support.workbook_api_fixtures import create_workbook_file, upload
from tests.test_workbook_writebacks import StubWritebackGenerator


def test_range_proposal_expands_and_reports_formula_impact() -> None:
    app.dependency_overrides[get_writeback_generator] = lambda: StubWritebackGenerator(
        "B2:C3", 0
    )
    try:
        response = TestClient(app).post(
            "/api/v1/workbooks/writeback-proposals",
            data={"instruction": "매출현황 B2:C3 범위를 0으로 수정해줘"},
            files=upload("sales.xlsx", create_workbook_file()),
        )
    finally:
        app.dependency_overrides.clear()
    payload = response.json()
    assert payload["status"] == "ready"
    assert [item["reference"] for item in payload["changes"]] == [
        "B2", "C2", "B3", "C3"
    ]
    assert payload["changes"][0]["affected_cells"] == ["매출현황!D2"]
    assert payload["changes"][0]["risk_level"] == "medium"


def test_valid_range_cells_remain_proposed_when_formula_cell_is_excluded() -> None:
    app.dependency_overrides[get_writeback_generator] = lambda: StubWritebackGenerator(
        "B2:D2", 0
    )
    try:
        payload = TestClient(app).post(
            "/api/v1/workbooks/writeback-proposals",
            data={"instruction": "매출현황 B2:D2 범위를 0으로 수정해줘"},
            files=upload("sales.xlsx", create_workbook_file()),
        ).json()
    finally:
        app.dependency_overrides.clear()
    assert payload["status"] == "ready"
    assert [item["reference"] for item in payload["changes"]] == ["B2", "C2"]
    assert "수식 셀" in payload["risks"][0]
    assert "적용 가능한 2개 셀" in payload["limitations"][0]


def test_explicit_safe_formula_can_be_proposed_and_applied() -> None:
    formula = "=SUM(B2:C2)*2"
    app.dependency_overrides[get_writeback_generator] = lambda: StubWritebackGenerator(
        "D2", formula
    )
    try:
        proposed = TestClient(app).post(
            "/api/v1/workbooks/writeback-proposals",
            data={"instruction": f"매출현황 D2 수식을 {formula}로 변경해줘"},
            files=upload("sales.xlsx", create_workbook_file()),
        ).json()
    finally:
        app.dependency_overrides.clear()
    assert proposed["status"] == "ready"
    assert proposed["changes"][0]["change_type"] == "formula"
    response = TestClient(app).post(
        "/api/v1/workbooks/writebacks/apply",
        data={"changes": json.dumps(proposed["changes"], ensure_ascii=False)},
        files=upload("sales.xlsx", create_workbook_file()),
    )
    with ZipFile(BytesIO(response.content)) as package:
        modified = package.read("workbook.xlsx")
    result = load_workbook(BytesIO(modified), data_only=False)
    try:
        assert result["매출현황"]["D2"].value == formula
    finally:
        result.close()


def test_clear_cell_is_applied_without_losing_style() -> None:
    original = create_workbook_file()
    changes = json.dumps(
        [{"sheet_name": "매출현황", "reference": "B3", "old_value": 5,
          "new_value": None, "reason": "값 삭제", "change_type": "clear",
          "value_type": "blank"}], ensure_ascii=False,
    )
    response = TestClient(app).post(
        "/api/v1/workbooks/writebacks/apply",
        data={"changes": changes}, files=upload("sales.xlsx", original),
    )
    with ZipFile(BytesIO(response.content)) as package:
        modified = package.read("workbook.xlsx")
    before = load_workbook(BytesIO(original), data_only=False)
    after = load_workbook(BytesIO(modified), data_only=False)
    try:
        assert after["매출현황"]["B3"].value is None
        assert before["매출현황"]["B3"].style_id == after["매출현황"]["B3"].style_id
    finally:
        before.close()
        after.close()


def test_date_change_preserves_excel_date_type_and_format() -> None:
    workbook = Workbook()
    workbook.active["A1"] = date(2026, 9, 1)
    workbook.active["A1"].number_format = "yyyy-mm-dd"
    source = BytesIO()
    workbook.save(source)
    workbook.close()
    changes = json.dumps(
        [{"sheet_name": "Sheet", "reference": "A1", "old_value": "2026-09-01",
          "new_value": "2026-09-30", "reason": "마감일 변경",
          "change_type": "value", "value_type": "date"}],
    )
    response = TestClient(app).post(
        "/api/v1/workbooks/writebacks/apply", data={"changes": changes},
        files=upload("schedule.xlsx", source.getvalue()),
    )
    with ZipFile(BytesIO(response.content)) as package:
        modified = package.read("workbook.xlsx")
    result = load_workbook(BytesIO(modified), data_only=False)
    try:
        assert result.active["A1"].value == datetime(2026, 9, 30)
        assert result.active["A1"].number_format == "yyyy-mm-dd"
    finally:
        result.close()
