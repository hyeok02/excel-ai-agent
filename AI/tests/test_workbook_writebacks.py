import json
from io import BytesIO
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.agent.writeback.models import WritebackChangeDraft, WritebackProposalDraft
from app.api.workbook_writebacks import get_writeback_generator
from app.main import app
from tests.support.workbook_api_fixtures import create_workbook_file, upload


class StubWritebackGenerator:
    def __init__(self, reference: str = "B2") -> None:
        self.reference = reference

    async def generate(self, instruction, filename, context):
        return WritebackProposalDraft(
            summary="1월 노트북 매출 값을 변경합니다.",
            changes=[
                WritebackChangeDraft(
                    sheet_name="매출현황",
                    reference=self.reference,
                    new_value=12,
                    reason="사용자가 정정 값을 명시했습니다.",
                )
            ],
        )


def test_proposal_returns_verified_old_and_new_values() -> None:
    app.dependency_overrides[get_writeback_generator] = lambda: StubWritebackGenerator()
    try:
        response = TestClient(app).post(
            "/api/v1/workbooks/writeback-proposals",
            data={"instruction": "매출현황 B2를 12로 수정해줘"},
            files=upload("sales.xlsx", create_workbook_file()),
        )
    finally:
        app.dependency_overrides.clear()
    assert response.status_code == 200
    assert response.json()["status"] == "ready"
    assert response.json()["changes"][0]["old_value"] == 10


def test_formula_target_is_blocked() -> None:
    app.dependency_overrides[get_writeback_generator] = lambda: StubWritebackGenerator("D2")
    try:
        response = TestClient(app).post(
            "/api/v1/workbooks/writeback-proposals",
            data={"instruction": "합계를 12로 바꿔줘"},
            files=upload("sales.xlsx", create_workbook_file()),
        )
    finally:
        app.dependency_overrides.clear()
    assert response.status_code == 200
    assert response.json()["status"] == "blocked"
    assert response.json()["changes"] == []
    assert "수식 셀" in response.json()["risks"][0]


def test_apply_returns_verified_copy_and_preserves_formula_and_style() -> None:
    original = create_workbook_file()
    changes = json.dumps(
        [{"sheet_name": "매출현황", "reference": "B2", "old_value": 10,
          "new_value": 12, "reason": "정정"}],
        ensure_ascii=False,
    )
    response = TestClient(app).post(
        "/api/v1/workbooks/writebacks/apply",
        data={"changes": changes},
        files=upload("sales.xlsx", original),
    )
    assert response.status_code == 200
    with ZipFile(BytesIO(response.content)) as package:
        modified = package.read("workbook.xlsx")
        manifest = json.loads(package.read("manifest.json"))
    before = load_workbook(BytesIO(original), data_only=False)
    after = load_workbook(BytesIO(modified), data_only=False)
    try:
        assert before["매출현황"]["B2"].value == 10
        assert after["매출현황"]["B2"].value == 12
        assert after["매출현황"]["D2"].value == "=SUM(B2:C2)"
        assert before["매출현황"]["B2"].style_id == after["매출현황"]["B2"].style_id
    finally:
        before.close()
        after.close()
    assert manifest["verified"] is True
    assert {item["name"] for item in manifest["checks"]} >= {
        "formulas", "styles", "unchanged_parts", "excel_extensions", "macros"
    }
